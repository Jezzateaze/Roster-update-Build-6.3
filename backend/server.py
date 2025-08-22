from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import JSONResponse, Response
from pymongo import MongoClient
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, time, timedelta
import os
import uuid
import hashlib
import secrets
from enum import Enum
import asyncio
import tempfile
from pathlib import Path

# OCR-related imports
import cv2
import numpy as np
import pytesseract
from pdf2image import convert_from_path
from PIL import Image, ImageEnhance
import pillow_heif
import magic
import logging

# Export-related imports
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import xlsxwriter
import csv
import io
from fastapi.responses import StreamingResponse

# Email notification imports
import aiosmtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from jinja2 import Template

# Database setup
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "shift_roster_db")

client = MongoClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="Shift Roster & Pay Calculator")

# CORS setup
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Email Configuration
EMAIL_CONFIG = {
    "smtp_server": "smtp.gmail.com",
    "smtp_port": 587,
    "use_tls": True,
    "admin_email": "jeremy.tomlinson88@gmail.com",
    "from_email": "noreply@workforce-management.com",  # This should be configured with proper SMTP credentials
    "from_name": "Workforce Management System"
}

# Enums
class PayMode(str, Enum):
    DEFAULT = "default"
    SCHADS = "schads"

class ShiftType(str, Enum):
    WEEKDAY_DAY = "weekday_day"
    WEEKDAY_EVENING = "weekday_evening"
    WEEKDAY_NIGHT = "weekday_night"
    SATURDAY = "saturday"
    SUNDAY = "sunday"
    PUBLIC_HOLIDAY = "public_holiday"
    SLEEPOVER = "sleepover"

class UserRole(str, Enum):
    ADMIN = "admin"
    SUPERVISOR = "supervisor"
    STAFF = "staff"

class FundingPeriod(str, Enum):
    MONTHLY = "monthly"
    QUARTERLY = "3_monthly"
    ANNUALLY = "annually"

class PlanManagement(str, Enum):
    SELF_MANAGED = "self_managed"
    PLAN_MANAGED = "plan_managed"
    NDIA_MANAGED = "ndia_managed"

# Client Profile Models
class EmergencyContact(BaseModel):
    name: str
    relationship: str
    mobile: str
    address: str

class PlanManagerDetails(BaseModel):
    provider_name: str
    contact_person: str
    phone: str
    email: str

class NDISFundingCategory(BaseModel):
    category_name: str
    total_amount: float
    funding_period: FundingPeriod
    description: str
    supports: List[Dict[str, Any]] = []
    spent_amount: float = 0.0
    remaining_amount: float = 0.0

class NDISPlan(BaseModel):
    plan_type: str  # e.g., "PACE"
    ndis_number: str
    plan_start_date: str
    plan_end_date: str
    plan_management: PlanManagement
    plan_manager: Optional[PlanManagerDetails] = None
    funding_categories: List[NDISFundingCategory] = []
    plan_document_path: Optional[str] = None
    created_at: datetime = datetime.now()
    updated_at: datetime = datetime.now()

class ClientBiography(BaseModel):
    strengths: str = ""
    living_arrangements: str = ""
    daily_life: str = ""
    goals: List[Dict[str, str]] = []  # Each goal has 'title', 'description', 'how_to_achieve'
    supports: List[Dict[str, str]] = []  # Each support has 'description', 'provider', 'frequency', 'type'
    additional_info: str = ""

class ClientProfile(BaseModel):
    id: str = None
    # Personal Information
    full_name: str
    date_of_birth: str
    age: Optional[int] = None
    sex: str
    disability_condition: str
    # Contact Information
    mobile: str
    address: str
    # Emergency Contacts
    emergency_contacts: List[EmergencyContact] = []
    # NDIS Plan
    current_ndis_plan: Optional[NDISPlan] = None
    # Biography Information
    biography: Optional[ClientBiography] = None
    # System fields
    created_at: datetime = datetime.now()
    updated_at: datetime = datetime.now()
    created_by: str = ""

# Pydantic models
class Staff(BaseModel):
    id: Optional[str] = None
    name: str
    active: bool = True
    created_at: Optional[datetime] = None

class ShiftTemplate(BaseModel):
    id: str
    name: str
    start_time: str
    end_time: str
    is_sleepover: bool = False
    day_of_week: int  # 0=Monday, 6=Sunday
    manual_shift_type: Optional[str] = None  # Override automatic shift type detection
    manual_hourly_rate: Optional[float] = None  # Override automatic hourly rate calculation
    allow_overlap: Optional[bool] = False  # Allow 2:1 shift overlapping

class RosterEntry(BaseModel):
    id: str
    date: str  # YYYY-MM-DD
    shift_template_id: str
    staff_id: Optional[str] = None
    staff_name: Optional[str] = None
    start_time: str
    end_time: str
    is_sleepover: bool = False
    is_public_holiday: bool = False
    manual_shift_type: Optional[str] = None  # Manual override for shift type
    manual_hourly_rate: Optional[float] = None  # Manual override for hourly rate
    manual_sleepover: Optional[bool] = None  # Manual override for sleepover status
    wake_hours: Optional[float] = None  # Additional wake hours beyond 2 hours
    hours_worked: float = 0.0
    base_pay: float = 0.0
    sleepover_allowance: float = 0.0
    total_pay: float = 0.0
    allow_overlap: Optional[bool] = False  # Allow this shift to overlap with others (for 2:1 shifts)
    
    # NDIS Charge Rate Fields (Client Billing)
    ndis_hourly_charge: float = 0.0  # NDIS hourly charge rate
    ndis_shift_charge: float = 0.0   # NDIS per-shift charge (for sleepovers)
    ndis_total_charge: float = 0.0   # Total NDIS charge for this shift
    ndis_line_item_code: Optional[str] = None  # NDIS line item code
    ndis_description: Optional[str] = None     # NDIS service description

class Settings(BaseModel):
    rates: Dict[str, float] = {
        "weekday_day": 42.00,
        "weekday_evening": 44.50,
        "weekday_night": 52.00,
        "saturday": 57.50,
        "sunday": 74.00,
        "public_holiday": 88.50,
        "sleepover_default": 175.00,
        "sleepover_schads": 286.56
    }
    
    # NDIS Invoice Charge Rates (Client Billing)
    ndis_charge_rates: Dict[str, Dict] = {
        "weekday_day": {
            "rate": 70.23,
            "line_item_code": "01_801_0115_1_1",
            "description": "Assistance in Supported Independent Living - Standard - Weekday Daytime",
            "time_range": "6am-8pm (Starts at/after 6:00am, ends at/before 8:00pm)"
        },
        "weekday_evening": {
            "rate": 77.38,
            "line_item_code": "01_802_0115_1_1", 
            "description": "Assistance in Supported Independent Living - Standard - Weekday Evening",
            "time_range": "After 8pm (Starts after 8:00pm OR extends past 8:00pm)"
        },
        "weekday_night": {
            "rate": 78.81,
            "line_item_code": "01_803_0115_1_1",
            "description": "Assistance in Supported Independent Living - Standard - Weekday Night", 
            "time_range": "Overnight (Commences at/before midnight and finishes after midnight)"
        },
        "saturday": {
            "rate": 98.83,
            "line_item_code": "01_804_0115_1_1",
            "description": "Assistance in Supported Independent Living - Standard - Saturday",
            "time_range": "All hours on Saturday"
        },
        "sunday": {
            "rate": 122.59,
            "line_item_code": "01_805_0115_1_1", 
            "description": "Assistance in Supported Independent Living - Standard - Sunday",
            "time_range": "All hours on Sunday"
        },
        "public_holiday": {
            "rate": 150.10,
            "line_item_code": "01_806_0115_1_1",
            "description": "Assistance in Supported Independent Living - Standard - Public Holiday",
            "time_range": "All hours on public holidays"
        },
        "sleepover_default": {
            "rate": 286.56,
            "line_item_code": "01_832_0115_1_1",
            "description": "Assistance in Supported Independent Living - Night-Time Sleepover",
            "time_range": "8-hour sleepover period (includes up to 2 hours active support)",
            "unit": "per_shift"
        }
    }
    
    first_day_of_week: str = "monday"  # "monday" or "sunday"
    pay_mode: str = "default"  # "default" or "schads"
    time_format: str = "24hr"  # "12hr" or "24hr"

class RosterTemplate(BaseModel):
    id: str
    name: str
    description: Optional[str] = None
    created_at: datetime = None
    is_active: bool = True
    template_data: Dict[str, List[Dict]] = {}  # day_of_week (as string) -> list of shift templates
    enable_2_1_shift: Optional[bool] = False  # Allow multiple staff on same shift time
    allow_overlap_override: Optional[bool] = False  # Allow override of overlap detection
    prevent_duplicate_unassigned: Optional[bool] = True  # Prevent duplicate entries if unassigned
    allow_different_staff_only: Optional[bool] = True  # Allow duplicates only if different staff assigned

class DayTemplate(BaseModel):
    id: str
    name: str
    description: Optional[str] = None
    day_of_week: int  # 0=Monday, 6=Sunday
    shifts: List[Dict] = []  # List of shift data (times, sleepover status)
    created_at: datetime = None
    is_active: bool = True

class CalendarEvent(BaseModel):
    id: str
    title: str
    description: Optional[str] = None
    date: str  # YYYY-MM-DD
    start_time: Optional[str] = None  # HH:MM (optional for all-day events)
    end_time: Optional[str] = None    # HH:MM (optional for all-day events)
    is_all_day: bool = False
    event_type: str = "appointment"  # appointment, meeting, task, reminder, personal
    priority: str = "medium"  # low, medium, high, urgent
    location: Optional[str] = None
    attendees: List[str] = []  # List of attendee names
    reminder_minutes: Optional[int] = None  # Minutes before event to remind
    is_completed: bool = False  # For tasks
    created_at: datetime = None
    is_active: bool = True

# Authentication Models
class User(BaseModel):
    id: str
    username: str
    pin_hash: str  # Hashed PIN
    role: UserRole = UserRole.STAFF
    staff_id: Optional[str] = None  # Link to staff record for staff users
    email: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    emergency_contact_name: Optional[str] = None
    emergency_contact_phone: Optional[str] = None
    profile_photo_url: Optional[str] = None
    date_of_birth: Optional[str] = None  # YYYY-MM-DD
    hire_date: Optional[str] = None      # YYYY-MM-DD
    hourly_rate: Optional[float] = None
    is_first_login: bool = True
    is_active: bool = True
    created_at: datetime = None
    last_login: Optional[datetime] = None

class LoginRequest(BaseModel):
    username: str
    pin: str

class ChangePinRequest(BaseModel):
    current_pin: str
    new_pin: str

class ResetPinRequest(BaseModel):
    username: str
    email: str

class Session(BaseModel):
    id: str
    user_id: str
    token: str
    created_at: datetime
    expires_at: datetime
    is_active: bool = True

# New models for Shift & Staff Availability System
class AvailabilityType(str, Enum):
    AVAILABLE = "available"
    UNAVAILABLE = "unavailable" 
    TIME_OFF_REQUEST = "time_off_request"
    PREFERRED_SHIFTS = "preferred_shifts"

class RequestStatus(str, Enum):
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"
    CANCELLED = "cancelled"

class NotificationType(str, Enum):
    SHIFT_REQUEST_APPROVED = "shift_request_approved"
    SHIFT_REQUEST_REJECTED = "shift_request_rejected"
    AVAILABILITY_CONFLICT = "availability_conflict"
    GENERAL = "general"

class ShiftRequest(BaseModel):
    id: Optional[str] = None
    roster_entry_id: str  # ID of the unassigned shift
    staff_id: str
    staff_name: str
    request_date: datetime
    status: RequestStatus = RequestStatus.PENDING
    notes: Optional[str] = None
    admin_notes: Optional[str] = None
    approved_by: Optional[str] = None
    approved_date: Optional[datetime] = None
    created_at: Optional[datetime] = None

class StaffAvailability(BaseModel):
    id: Optional[str] = None
    staff_id: str
    staff_name: str
    availability_type: AvailabilityType
    date_from: Optional[str] = None  # YYYY-MM-DD for specific dates
    date_to: Optional[str] = None    # YYYY-MM-DD for date ranges
    day_of_week: Optional[int] = None  # 0-6 for recurring weekly availability
    start_time: Optional[str] = None   # HH:MM for time-specific availability
    end_time: Optional[str] = None     # HH:MM for time-specific availability
    is_recurring: bool = False         # Whether this applies every week
    notes: Optional[str] = None
    created_at: Optional[datetime] = None
    is_active: bool = True

# OCR Document Processing Models
class OCRTaskStatus(str, Enum):
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"

class DocumentType(str, Enum):
    PDF = "pdf"
    IMAGE = "image"

class OCRTask(BaseModel):
    id: Optional[str] = None
    client_id: Optional[str] = None  # Link to client profile
    filename: str
    file_type: str
    status: OCRTaskStatus = OCRTaskStatus.PROCESSING
    progress: int = 0
    result: Optional[Dict[str, Any]] = None
    error: Optional[str] = None
    extracted_text: Optional[str] = None
    extracted_data: Optional[Dict[str, Any]] = None  # Parsed NDIS plan data
    created_at: datetime = datetime.now()
    completed_at: Optional[datetime] = None
    created_by: str = ""

class ExtractedNDISData(BaseModel):
    full_name: Optional[str] = None
    date_of_birth: Optional[str] = None
    ndis_number: Optional[str] = None
    plan_start_date: Optional[str] = None
    plan_end_date: Optional[str] = None
    funding_categories: List[Dict[str, Any]] = []
    disability_condition: Optional[str] = None
    address: Optional[str] = None
    mobile: Optional[str] = None
    confidence_score: float = 0.0

# Notification Models
class Notification(BaseModel):
    id: Optional[str] = None
    user_id: str
    notification_type: NotificationType
    title: str
    message: str
    related_id: Optional[str] = None  # ID of related shift request, availability, etc.
    is_read: bool = False
    created_at: Optional[datetime] = None

# Authentication helper functions
def hash_pin(pin: str) -> str:
    """Hash a PIN using SHA-256"""
    return hashlib.sha256(pin.encode()).hexdigest()

def verify_pin(pin: str, pin_hash: str) -> bool:
    """Verify a PIN against its hash"""
    return hash_pin(pin) == pin_hash

def generate_token() -> str:
    """Generate a secure random token"""
    return secrets.token_urlsafe(32)

def create_admin_user():
    """Create the default admin user if it doesn't exist"""
    admin_user = db.users.find_one({"username": "Admin"})
    if not admin_user:
        admin_data = {
            "id": str(uuid.uuid4()),
            "username": "Admin",
            "pin_hash": hash_pin("0000"),  # Default PIN: 0000
            "role": "admin",
            "email": "jeremy.tomlinson88@gmail.com",
            "first_name": "Administrator",
            "is_first_login": True,
            "is_active": True,
            "created_at": datetime.utcnow()
        }
        db.users.insert_one(admin_data)
        print("✅ Admin user created with default PIN: 0000")

def send_reset_email(email: str, temp_pin: str):
    """Send password reset email (simplified version)"""
    # In a production environment, you would configure proper SMTP settings
    # For now, we'll just print the reset PIN to console
    print(f"🔐 RESET PIN for {email}: {temp_pin}")
    print("📧 In production, this would be sent via email")

# Initialize admin user on startup
create_admin_user()

# OCR Configuration and Processing
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Initialize HEIF plugin for Pillow
pillow_heif.register_heif_opener()

ALLOWED_EXTENSIONS = {
    'image/jpeg', 'image/jpg', 'image/png', 'image/tiff', 'image/bmp',
    'image/heif', 'image/heic', 'image/heif-sequence', 'image/heic-sequence',
    'application/pdf'
}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

# In-memory storage for OCR results (use Redis or database in production)
ocr_results: Dict[str, Dict[str, Any]] = {}

class NDISOCRProcessor:
    """OCR processor specialized for NDIS plan documents"""
    
    def __init__(self):
        # Tesseract configuration optimized for forms and structured documents
        self.tesseract_config = r'--oem 3 --psm 6 -l eng'
        self.dpi = 300  # Optimal DPI for OCR accuracy
        self.logger = logging.getLogger(__name__)
        
    def preprocess_image(self, image: np.ndarray) -> np.ndarray:
        """Apply preprocessing techniques optimized for NDIS documents"""
        # Convert to grayscale
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image
            
        # Apply denoising
        denoised = cv2.fastNlMeansDenoising(gray)
        
        # Enhance contrast using CLAHE
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        enhanced = clahe.apply(denoised)
        
        # Apply adaptive thresholding
        binary = cv2.adaptiveThreshold(
            enhanced, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
            cv2.THRESH_BINARY, 11, 2
        )
        
        # Deskew the image
        deskewed = self.correct_skew(binary)
        
        return deskewed
    
    def correct_skew(self, image: np.ndarray) -> np.ndarray:
        """Correct document skew for better OCR accuracy"""
        coords = np.column_stack(np.where(image > 0))
        if len(coords) == 0:
            return image
            
        angle = cv2.minAreaRect(coords)[-1]
        
        # Normalize angle
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle
            
        # Only correct if angle is significant
        if abs(angle) > 0.5:
            (h, w) = image.shape[:2]
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, angle, 1.0)
            rotated = cv2.warpAffine(
                image, M, (w, h),
                flags=cv2.INTER_CUBIC,
                borderMode=cv2.BORDER_REPLICATE
            )
            return rotated
        
        return image
    
    def process_pdf(self, pdf_path: Path) -> List[Dict]:
        """Convert PDF pages to images and extract text"""
        try:
            self.logger.info(f"📄 Starting PDF processing: {pdf_path}")
            
            # Check if file exists and is accessible
            if not pdf_path.exists():
                raise FileNotFoundError(f"PDF file not found: {pdf_path}")
            
            # Log file size for debugging
            file_size = pdf_path.stat().st_size
            self.logger.info(f"📊 PDF file size: {file_size} bytes")
            
            # Convert PDF pages to images
            self.logger.info(f"🔄 Converting PDF pages to images (DPI: {self.dpi})")
            pages = convert_from_path(
                str(pdf_path), 
                dpi=self.dpi,
                first_page=1,
                last_page=None,
                thread_count=1
            )
            
            self.logger.info(f"✅ Successfully converted PDF to {len(pages)} pages")
            
            extracted_texts = []
            
            for page_num, page in enumerate(pages, 1):
                self.logger.info(f"🔍 Processing page {page_num}/{len(pages)}")
                
                # Convert PIL image to numpy array
                page_array = np.array(page)
                
                # Preprocess the image
                processed_image = self.preprocess_image(page_array)
                
                # Extract text using Tesseract
                text = pytesseract.image_to_string(
                    processed_image, 
                    config=self.tesseract_config
                )
                
                word_count = len(text.split()) if text.strip() else 0
                self.logger.info(f"📝 Page {page_num}: extracted {word_count} words")
                
                extracted_texts.append({
                    'page': page_num,
                    'text': text.strip(),
                    'word_count': word_count
                })
                
            total_words = sum(page['word_count'] for page in extracted_texts)
            self.logger.info(f"✅ PDF processing complete: {len(pages)} pages, {total_words} total words")
            
            return extracted_texts
            
        except FileNotFoundError as e:
            self.logger.error(f"❌ PDF file not found: {str(e)}")
            raise HTTPException(status_code=404, detail=f"PDF file not found: {str(e)}")
        except Exception as e:
            error_msg = str(e)
            self.logger.error(f"❌ Error processing PDF {pdf_path}: {error_msg}")
            
            # More specific error messages
            if "poppler" in error_msg.lower():
                raise HTTPException(status_code=500, detail="PDF processing tools not available. Please contact support.")
            elif "permission" in error_msg.lower():
                raise HTTPException(status_code=500, detail="Permission error accessing PDF file.")
            elif "corrupt" in error_msg.lower() or "damaged" in error_msg.lower():
                raise HTTPException(status_code=400, detail="PDF file appears to be corrupted or damaged.")
            else:
                raise HTTPException(status_code=500, detail=f"PDF processing failed: {error_msg}")
    
    def process_image(self, image_path: Path) -> Dict[str, Any]:
        """Process image file and extract text"""
        try:
            # Check if it's a HEIF/HEIC file
            file_extension = image_path.suffix.lower()
            is_heif = file_extension in ['.heif', '.heic']
            
            if is_heif:
                # Load HEIF/HEIC image using Pillow with HEIF support
                self.logger.info(f"Loading HEIF/HEIC image: {image_path}")
                pil_image = Image.open(str(image_path))
                # Convert to RGB if necessary
                if pil_image.mode != 'RGB':
                    pil_image = pil_image.convert('RGB')
                # Convert PIL image to numpy array for OpenCV
                image = np.array(pil_image)
                # Convert RGB to BGR for OpenCV
                image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
                self.logger.info(f"Successfully loaded HEIF image with dimensions: {image.shape}")
            else:
                # Load regular image formats using OpenCV
                image = cv2.imread(str(image_path))
                if image is None:
                    # Fallback: try loading with Pillow and convert
                    pil_image = Image.open(str(image_path))
                    if pil_image.mode != 'RGB':
                        pil_image = pil_image.convert('RGB')
                    image = np.array(pil_image)
                    image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
            
            if image is None:
                raise ValueError("Could not load image")
            
            # Preprocess the image
            processed_image = self.preprocess_image(image)
            
            # Extract text using Tesseract
            text = pytesseract.image_to_string(
                processed_image, 
                config=self.tesseract_config
            )
            
            return {
                'text': text.strip(),
                'word_count': len(text.split()) if text.strip() else 0,
                'confidence': self.get_text_confidence(processed_image),
                'format': 'heif' if is_heif else 'standard_image'
            }
            
        except Exception as e:
            self.logger.error(f"Error processing image {image_path}: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Image processing failed: {str(e)}")
    
    def get_text_confidence(self, image: np.ndarray) -> float:
        """Calculate OCR confidence score"""
        try:
            data = pytesseract.image_to_data(
                image, 
                output_type=pytesseract.Output.DICT,
                config=self.tesseract_config
            )
            
            confidences = [int(conf) for conf in data['conf'] if int(conf) > 0]
            
            if confidences:
                return sum(confidences) / len(confidences)
            return 0.0
            
        except Exception:
            return 0.0
    
    def parse_ndis_plan_text(self, text: str) -> ExtractedNDISData:
        """Parse NDIS plan information from extracted text"""
        extracted_data = ExtractedNDISData()
        
        # Split text into lines for parsing
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        # Common patterns for NDIS plan documents
        import re
        
        try:
            # Name patterns
            name_patterns = [
                r'(?:participant|name|full name|client name):\s*(.+?)(?:\n|$)',
                r'name\s*[:\-]\s*(.+?)(?:\n|$)',
                r'participant\s+name\s*[:\-]\s*(.+?)(?:\n|$)',
            ]
            
            for pattern in name_patterns:
                match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                if match and not extracted_data.full_name:
                    extracted_data.full_name = match.group(1).strip()
                    break
            
            # NDIS number patterns
            ndis_patterns = [
                r'(?:ndis|participant)\s*(?:number|id|#):\s*(\d{9})',
                r'(?:ndis|participant)\s*(?:number|id|#)\s*[:\-]\s*(\d{9})',
                r'participant\s+id\s*[:\-]\s*(\d{9})',
            ]
            
            for pattern in ndis_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match and not extracted_data.ndis_number:
                    extracted_data.ndis_number = match.group(1).strip()
                    break
            
            # Date of birth patterns
            dob_patterns = [
                r'(?:date of birth|dob|born):\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
                r'(?:date of birth|dob|born)\s*[:\-]\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
            ]
            
            for pattern in dob_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match and not extracted_data.date_of_birth:
                    extracted_data.date_of_birth = match.group(1).strip()
                    break
            
            # Plan period patterns
            plan_start_patterns = [
                r'plan\s+(?:starts?|from):\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
                r'start\s+date:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
                r'plan\s+period:\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
            ]
            
            for pattern in plan_start_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match and not extracted_data.plan_start_date:
                    extracted_data.plan_start_date = match.group(1).strip()
                    break
            
            plan_end_patterns = [
                r'(?:plan\s+(?:ends?|to)|end\s+date):\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
                r'plan\s+period:.*?to\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
            ]
            
            for pattern in plan_end_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match and not extracted_data.plan_end_date:
                    extracted_data.plan_end_date = match.group(1).strip()
                    break
            
            # Disability condition patterns
            disability_patterns = [
                r'(?:primary disability|disability|condition):\s*(.+?)(?:\n|$)',
                r'(?:diagnosis|condition|disability)\s*[:\-]\s*(.+?)(?:\n|$)',
            ]
            
            for pattern in disability_patterns:
                match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                if match and not extracted_data.disability_condition:
                    extracted_data.disability_condition = match.group(1).strip()
                    break
            
            # Address patterns
            address_patterns = [
                r'address:\s*(.+?)(?:\n|$)',
                r'address\s*[:\-]\s*(.+?)(?:\n|$)',
            ]
            
            for pattern in address_patterns:
                match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                if match and not extracted_data.address:
                    extracted_data.address = match.group(1).strip()
                    break
            
            # Mobile phone patterns
            mobile_patterns = [
                r'(?:mobile|phone|contact):\s*(\d{10}|\(\d{2}\)\s*\d{4}\s*\d{4})',
                r'(?:mobile|phone)\s*[:\-]\s*(\d{10}|\(\d{2}\)\s*\d{4}\s*\d{4})',
            ]
            
            for pattern in mobile_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match and not extracted_data.mobile:
                    extracted_data.mobile = match.group(1).strip()
                    break
            
            # Calculate confidence based on how many fields we found
            fields_found = sum(1 for field in [
                extracted_data.full_name,
                extracted_data.ndis_number,
                extracted_data.date_of_birth,
                extracted_data.plan_start_date,
                extracted_data.plan_end_date
            ] if field)
            
            extracted_data.confidence_score = (fields_found / 5.0) * 100
            
        except Exception as e:
            self.logger.error(f"Error parsing NDIS plan text: {str(e)}")
        
        return extracted_data

# OCR processor instance
ocr_processor = NDISOCRProcessor()

async def validate_and_save_file(file: UploadFile) -> Path:
    """Validate and save uploaded file"""
    # Check file size
    file_content = await file.read()
    if len(file_content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large (max 50MB)")
    
    # Validate file type - enhanced for mobile/iOS compatibility
    file_type = magic.from_buffer(file_content[:1024], mime=True)
    file_extension = Path(file.filename).suffix.lower() if file.filename else ''
    
    logging.info(f"File validation: {file.filename}, MIME: {file_type}, Extension: {file_extension}")
    
    # Enhanced PDF validation for iOS/mobile compatibility
    is_pdf_by_extension = file_extension in ['.pdf']
    is_pdf_by_mime = file_type in ['application/pdf']
    is_pdf_by_content = file_content[:4] == b'%PDF'  # PDF magic bytes
    
    # iOS Safari sometimes sends PDFs with generic MIME types
    ios_pdf_mimes = [
        'application/octet-stream',  # Common iOS Safari behavior
        'application/binary',
        'application/download',
        'text/plain'  # Sometimes iOS sends this for PDFs
    ]
    is_pdf_ios = file_type in ios_pdf_mimes and is_pdf_by_extension and is_pdf_by_content
    
    # HEIF/HEIC validation (already implemented)
    is_heif_by_extension = file_extension in ['.heif', '.heic']
    is_heif_by_mime = file_type in ['image/heif', 'image/heic', 'image/heif-sequence', 'image/heic-sequence']
    
    # Check if file is valid
    valid_file = (
        file_type in ALLOWED_EXTENSIONS or  # Standard MIME types
        is_pdf_by_extension and (is_pdf_by_mime or is_pdf_by_content or is_pdf_ios) or  # PDF validation
        is_heif_by_extension or is_heif_by_mime  # HEIF validation
    )
    
    if not valid_file:
        error_detail = (
            f"Unsupported file type: {file_type} (extension: {file_extension}). "
            f"Supported formats: PDF, JPG, PNG, TIFF, BMP, HEIF/HEIC. "
        )
        
        # Add specific guidance for PDF issues
        if file_extension == '.pdf':
            error_detail += (
                "PDF file detected by extension but content validation failed. "
                "Please ensure the file is a valid PDF document."
            )
        
        logging.error(f"File validation failed: {error_detail}")
        raise HTTPException(status_code=400, detail=error_detail)
    
    # Save file with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_filename = f"{timestamp}_{file.filename}"
    file_path = UPLOAD_DIR / safe_filename
    
    with open(file_path, 'wb') as buffer:
        buffer.write(file_content)
    
    logging.info(f"Successfully saved file: {safe_filename} (MIME: {file_type}, Extension: {file_extension}, PDF checks: mime={is_pdf_by_mime}, content={is_pdf_by_content}, ios={is_pdf_ios})")
    
    return file_path

async def process_pdf_async(file_path: Path, task_id: str) -> Dict[str, Any]:
    """Asynchronously process PDF file"""
    def run_ocr():
        return ocr_processor.process_pdf(file_path)
    
    # Update progress
    if task_id in ocr_results:
        ocr_results[task_id]['progress'] = 50
    
    # Run OCR in thread pool to avoid blocking
    loop = asyncio.get_event_loop()
    pages_text = await loop.run_in_executor(None, run_ocr)
    
    # Combine all text for parsing
    combined_text = "\n".join(page['text'] for page in pages_text)
    
    return {
        'document_type': 'pdf',
        'total_pages': len(pages_text),
        'pages': pages_text,
        'combined_text': combined_text,
        'total_words': sum(page['word_count'] for page in pages_text),
        'extracted_at': datetime.now().isoformat()
    }

async def process_image_async(file_path: Path, task_id: str) -> Dict[str, Any]:
    """Asynchronously process image file"""
    def run_ocr():
        return ocr_processor.process_image(file_path)
    
    # Update progress
    if task_id in ocr_results:
        ocr_results[task_id]['progress'] = 50
    
    # Run OCR in thread pool
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(None, run_ocr)
    
    return {
        'document_type': 'image',
        'text': result['text'],
        'word_count': result['word_count'],
        'confidence': result['confidence'],
        'extracted_at': datetime.now().isoformat()
    }

# Pay calculation functions
def determine_shift_type_with_context(date_str: str, start_time: str, end_time: str, is_public_holiday: bool, is_post_midnight_segment: bool = False) -> ShiftType:
    """Determine the shift type with context for cross-midnight calculations"""
    
    if is_public_holiday:
        return ShiftType.PUBLIC_HOLIDAY
    
    # Parse date and get day of week
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    day_of_week = date_obj.weekday()  # 0=Monday, 1=Tuesday, ..., 6=Sunday
    
    # Weekend rates override time-based logic
    if day_of_week == 5:  # Saturday
        return ShiftType.SATURDAY
    elif day_of_week == 6:  # Sunday
        return ShiftType.SUNDAY
    
    # For weekdays (Monday-Friday), check time ranges
    start_hour = int(start_time.split(":")[0])
    start_min = int(start_time.split(":")[1])
    end_hour = int(end_time.split(":")[0])
    end_min = int(end_time.split(":")[1])
    
    start_minutes = start_hour * 60 + start_min
    end_minutes = end_hour * 60 + end_min
    
    # Handle overnight shifts
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60
    
    # Special handling for post-midnight segments of cross-midnight shifts
    if is_post_midnight_segment:
        # For post-midnight segments (00:01-XX:XX), classify based on the end time
        # Since this segment starts at midnight, we determine the rate based on 
        # what type of shift it would be considered ending at that time
        
        # If the segment ends at 6am or later, it's entering day shift territory
        if end_hour >= 6:
            return ShiftType.WEEKDAY_DAY
        # If ending before 6am, it's still night shift
        else:
            return ShiftType.WEEKDAY_NIGHT
    
    # Special handling for shifts starting exactly at midnight (00:00)
    if start_hour == 0 and start_min == 0:
        # Shifts starting at midnight should be classified based on end time
        if end_hour >= 6:
            return ShiftType.WEEKDAY_DAY
        else:
            return ShiftType.WEEKDAY_NIGHT
    
    # Original logic for regular shifts and first segments of cross-midnight shifts
    # Night: starts before 6am OR ends after midnight
    if start_hour < 6 or end_minutes > 24 * 60:
        return ShiftType.WEEKDAY_NIGHT
    # Evening: starts at 8pm or later OR extends past 8pm (SCHADS: shifts ending AT 8pm are still Day shifts)
    elif start_hour >= 20 or end_minutes > 20 * 60:
        return ShiftType.WEEKDAY_EVENING
    # Day: everything else (6am-8pm range)
    else:
        return ShiftType.WEEKDAY_DAY

def determine_shift_type(date_str: str, start_time: str, end_time: str, is_public_holiday: bool) -> ShiftType:
    """Determine the shift type based on date and time - SIMPLIFIED LOGIC"""
    return determine_shift_type_with_context(date_str, start_time, end_time, is_public_holiday, False)

def calculate_hours_worked(start_time: str, end_time: str) -> float:
    """Calculate hours worked between start and end time"""
    start_hour, start_min = map(int, start_time.split(":"))
    end_hour, end_min = map(int, end_time.split(":"))
    
    start_minutes = start_hour * 60 + start_min
    end_minutes = end_hour * 60 + end_min
    
    # Handle overnight shifts
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60
    
    total_minutes = end_minutes - start_minutes
    return total_minutes / 60.0

def check_shift_overlap(date_str: str, start_time: str, end_time: str, exclude_id: Optional[str] = None, shift_name: Optional[str] = None) -> bool:
    """Check if a shift overlaps with existing shifts on the same date
    
    Args:
        date_str: Date in YYYY-MM-DD format
        start_time: Start time in HH:MM format
        end_time: End time in HH:MM format
        exclude_id: ID of shift to exclude from overlap check (for updates)
        shift_name: Name of the shift being checked (to allow 2:1 overlaps)
    
    Returns:
        True if overlap detected and should be prevented, False if overlap is allowed
    """
    existing_shifts = list(db.roster.find({"date": date_str}))
    
    if exclude_id:
        existing_shifts = [s for s in existing_shifts if s.get("id") != exclude_id]
    
    if not existing_shifts:
        return False
    
    # Convert times to minutes for comparison
    def time_to_minutes(time_str: str) -> int:
        hours, minutes = map(int, time_str.split(':'))
        return hours * 60 + minutes
    
    new_start = time_to_minutes(start_time)
    new_end = time_to_minutes(end_time)
    
    # Handle overnight shift
    if new_end <= new_start:
        new_end += 24 * 60
    
    for shift in existing_shifts:
        existing_start = time_to_minutes(shift["start_time"])
        existing_end = time_to_minutes(shift["end_time"])
        
        # Handle overnight shift
        if existing_end <= existing_start:
            existing_end += 24 * 60
        
        # Check for overlap
        if (new_start < existing_end and new_end > existing_start):
            # Check if either shift has "2:1" in the name - if so, allow overlap
            existing_shift_name = ""
            
            # Try to get shift name from template or roster entry
            if shift.get("shift_template_id"):
                template = db.shift_templates.find_one({"id": shift["shift_template_id"]})
                if template:
                    existing_shift_name = template.get("name", "")
            
            # Also check if the shift itself has a name field
            if not existing_shift_name and shift.get("name"):
                existing_shift_name = shift.get("name", "")
            
            # Check both current shift name and existing shift name for "2:1"
            current_has_2to1 = shift_name and "2:1" in shift_name.lower()
            existing_has_2to1 = "2:1" in existing_shift_name.lower()
            
            # Only allow overlap if BOTH shifts have "2:1" in their names
            # OR if the current shift being added/updated has "2:1" in its name
            if current_has_2to1:
                print(f"Allowing overlap for 2:1 shift - Current: {shift_name}, Existing: {existing_shift_name}")
                continue  # Allow this overlap because current shift is 2:1
            else:
                return True  # Prevent overlap - regular shifts cannot overlap with anything
    
    return False

def calculate_ndis_charges(roster_entry: RosterEntry, settings: Settings, shift_type: str) -> RosterEntry:
    """Calculate NDIS charges for client billing"""
    # Determine if this is a sleepover shift
    is_sleepover = roster_entry.manual_sleepover if roster_entry.manual_sleepover is not None else roster_entry.is_sleepover
    
    if is_sleepover:
        # Sleepover charges are per-shift basis
        sleepover_charge = settings.ndis_charge_rates["sleepover_default"]
        roster_entry.ndis_hourly_charge = 0.0  # Not applicable for sleepovers
        roster_entry.ndis_shift_charge = sleepover_charge["rate"]
        roster_entry.ndis_total_charge = sleepover_charge["rate"]
        roster_entry.ndis_line_item_code = sleepover_charge["line_item_code"]
        roster_entry.ndis_description = sleepover_charge["description"]
        
        # Add any additional wake hours beyond 2 hours at hourly rate
        wake_hours = roster_entry.wake_hours if roster_entry.wake_hours else 0
        extra_wake_hours = max(0, wake_hours - 2) if wake_hours > 2 else 0
        
        if extra_wake_hours > 0:
            # For extra wake hours beyond 2, use the appropriate hourly NDIS rate
            # based on the actual shift timing (not sleepover rate)
            # Determine the proper shift type for NDIS hourly billing
            if shift_type == "sleepover_default":
                # For sleepover extra wake hours, always use weekday_day rate
                # as per review request: 1 hour × $70.23 (weekday_day NDIS rate)
                # This is because the sleepover base rate already includes night premium
                from datetime import datetime
                try:
                    shift_date = datetime.strptime(roster_entry.date, "%Y-%m-%d")
                    day_of_week = shift_date.weekday()  # 0=Monday, 6=Sunday
                    
                    if roster_entry.is_public_holiday:
                        ndis_shift_type_key = "public_holiday"
                    elif day_of_week == 5:  # Saturday
                        ndis_shift_type_key = "saturday"
                    elif day_of_week == 6:  # Sunday
                        ndis_shift_type_key = "sunday"
                    else:
                        # For weekday sleepovers, always use weekday_day rate for extra wake hours
                        # This matches the expected calculation in the review request
                        ndis_shift_type_key = "weekday_day"
                        
                except Exception as e:
                    # Default to weekday_day for sleepover extra wake hours
                    # This ensures we use the correct rate as specified in review request
                    ndis_shift_type_key = "weekday_day"
            else:
                ndis_shift_type_key = shift_type.lower().replace('_', '_')
            
            if ndis_shift_type_key in settings.ndis_charge_rates:
                hourly_ndis_rate = settings.ndis_charge_rates[ndis_shift_type_key]["rate"]
                roster_entry.ndis_total_charge += extra_wake_hours * hourly_ndis_rate
    else:
        # Regular shift - hourly charges
        # Map shift_type to NDIS charge rate key
        shift_type_key = shift_type.lower().replace('_', '_')
        
        if shift_type_key in settings.ndis_charge_rates:
            ndis_rate_info = settings.ndis_charge_rates[shift_type_key]
            roster_entry.ndis_hourly_charge = ndis_rate_info["rate"]
            roster_entry.ndis_shift_charge = 0.0  # Not applicable for regular shifts
            roster_entry.ndis_total_charge = roster_entry.hours_worked * ndis_rate_info["rate"]
            roster_entry.ndis_line_item_code = ndis_rate_info["line_item_code"]
            roster_entry.ndis_description = ndis_rate_info["description"]
        else:
            # Default to weekday_day if shift type not found
            default_rate = settings.ndis_charge_rates["weekday_day"]
            roster_entry.ndis_hourly_charge = default_rate["rate"]
            roster_entry.ndis_shift_charge = 0.0
            roster_entry.ndis_total_charge = roster_entry.hours_worked * default_rate["rate"]
            roster_entry.ndis_line_item_code = default_rate["line_item_code"]
            roster_entry.ndis_description = default_rate["description"]
    
    return roster_entry

def calculate_cross_midnight_pay(roster_entry: RosterEntry, settings: Settings) -> RosterEntry:
    """Calculate pay for shifts that cross midnight, splitting by actual days"""
    from datetime import datetime, timedelta
    
    # Parse the shift date and times
    shift_date = datetime.strptime(roster_entry.date, "%Y-%m-%d")
    start_hour, start_min = map(int, roster_entry.start_time.split(":"))
    end_hour, end_min = map(int, roster_entry.end_time.split(":"))
    
    start_minutes = start_hour * 60 + start_min
    end_minutes = end_hour * 60 + end_min
    
    # Check if shift crosses midnight
    crosses_midnight = end_minutes <= start_minutes
    
    if not crosses_midnight:
        # Regular shift - use existing logic
        return calculate_pay_regular(roster_entry, settings)
    
    # Shift crosses midnight - split into segments
    is_sleepover = roster_entry.manual_sleepover if roster_entry.manual_sleepover is not None else roster_entry.is_sleepover
    
    if is_sleepover:
        # Sleepover shifts don't need splitting - use existing logic
        return calculate_pay_regular(roster_entry, settings)
    
    # Split the shift at midnight
    # Segment 1: start_time to 23:59 on the first day
    first_day_end_minutes = 24 * 60  # midnight
    first_day_duration_minutes = first_day_end_minutes - start_minutes
    first_day_hours = first_day_duration_minutes / 60.0
    
    # Segment 2: 00:00 to end_time on the second day
    second_day_start_minutes = 0  # midnight
    second_day_duration_minutes = end_minutes
    second_day_hours = second_day_duration_minutes / 60.0
    
    # Get the second day date
    second_day_date = shift_date + timedelta(days=1)
    second_day_date_str = second_day_date.strftime("%Y-%m-%d")
    
    # Calculate pay for first day segment
    first_day_shift_type = determine_shift_type(
        roster_entry.date, 
        roster_entry.start_time, 
        "23:59",
        roster_entry.is_public_holiday
    )
    first_day_rate = get_hourly_rate_for_shift_type(first_day_shift_type, settings)
    first_day_pay = first_day_hours * first_day_rate
    
    # Calculate pay for second day segment
    # Check if second day is a public holiday
    second_day_is_holiday = is_public_holiday_date(second_day_date_str, settings)
    
    second_day_shift_type = determine_shift_type_with_context(
        second_day_date_str,
        "00:01",  # Just after midnight
        roster_entry.end_time,
        second_day_is_holiday,
        is_post_midnight_segment=True  # This is the post-midnight segment
    )
    second_day_rate = get_hourly_rate_for_shift_type(second_day_shift_type, settings)
    second_day_pay = second_day_hours * second_day_rate
    
    # Set the calculated values
    roster_entry.hours_worked = first_day_hours + second_day_hours
    roster_entry.base_pay = first_day_pay + second_day_pay
    roster_entry.sleepover_allowance = 0
    roster_entry.total_pay = roster_entry.base_pay
    
    # Calculate NDIS charges - use the primary (first day) shift type
    if roster_entry.manual_shift_type:
        ndis_shift_type = roster_entry.manual_shift_type
    else:
        # Use first day shift type for NDIS (primary segment)
        if first_day_shift_type == ShiftType.PUBLIC_HOLIDAY:
            ndis_shift_type = "public_holiday"
        elif first_day_shift_type == ShiftType.SATURDAY:
            ndis_shift_type = "saturday"
        elif first_day_shift_type == ShiftType.SUNDAY:
            ndis_shift_type = "sunday"
        elif first_day_shift_type == ShiftType.WEEKDAY_EVENING:
            ndis_shift_type = "weekday_evening"
        elif first_day_shift_type == ShiftType.WEEKDAY_NIGHT:
            ndis_shift_type = "weekday_night"
        else:
            ndis_shift_type = "weekday_day"
    
    # Calculate NDIS charges
    roster_entry = calculate_ndis_charges(roster_entry, settings, ndis_shift_type)
    
    return roster_entry

def get_hourly_rate_for_shift_type(shift_type: ShiftType, settings: Settings) -> float:
    """Get hourly rate for a given shift type"""
    if shift_type == ShiftType.PUBLIC_HOLIDAY:
        return settings.rates["public_holiday"]
    elif shift_type == ShiftType.SATURDAY:
        return settings.rates["saturday"]
    elif shift_type == ShiftType.SUNDAY:
        return settings.rates["sunday"]
    elif shift_type == ShiftType.WEEKDAY_EVENING:
        return settings.rates["weekday_evening"]
    elif shift_type == ShiftType.WEEKDAY_NIGHT:
        return settings.rates["weekday_night"]
    else:
        return settings.rates["weekday_day"]

def is_public_holiday_date(date_str: str, settings: Settings) -> bool:
    """Check if a specific date is a public holiday"""
    # This is a simplified check - you may want to implement proper holiday checking
    # For now, returning False, but you can enhance this with actual holiday data
    return False

def calculate_pay_regular(roster_entry: RosterEntry, settings: Settings) -> RosterEntry:
    """Original calculate_pay logic for non-cross-midnight shifts"""
    hours = calculate_hours_worked(roster_entry.start_time, roster_entry.end_time)
    roster_entry.hours_worked = hours
    
    # Determine if this is a sleepover shift
    is_sleepover = roster_entry.manual_sleepover if roster_entry.manual_sleepover is not None else roster_entry.is_sleepover
    
    if is_sleepover:
        # Sleepover calculation: $175 flat rate includes 2 hours
        roster_entry.sleepover_allowance = 175.00  # Fixed $175 per night
        
        # Additional wake hours beyond 2 hours at applicable hourly rate
        wake_hours = roster_entry.wake_hours if roster_entry.wake_hours else 0
        extra_wake_hours = max(0, wake_hours - 2) if wake_hours > 2 else 0
        
        if extra_wake_hours > 0:
            # Get applicable hourly rate for extra wake time
            if roster_entry.manual_hourly_rate:
                hourly_rate = roster_entry.manual_hourly_rate
            else:
                # Determine rate based on shift type or manual override
                if roster_entry.manual_shift_type:
                    shift_type_map = {
                        "weekday_day": ShiftType.WEEKDAY_DAY,
                        "weekday_evening": ShiftType.WEEKDAY_EVENING,
                        "weekday_night": ShiftType.WEEKDAY_NIGHT,
                        "saturday": ShiftType.SATURDAY,
                        "sunday": ShiftType.SUNDAY,
                        "public_holiday": ShiftType.PUBLIC_HOLIDAY
                    }
                    shift_type = shift_type_map.get(roster_entry.manual_shift_type, ShiftType.WEEKDAY_DAY)
                else:
                    shift_type = determine_shift_type(
                        roster_entry.date, 
                        roster_entry.start_time, 
                        roster_entry.end_time,
                        roster_entry.is_public_holiday
                    )
                
                hourly_rate = get_hourly_rate_for_shift_type(shift_type, settings)
            
            roster_entry.base_pay = extra_wake_hours * hourly_rate
        else:
            roster_entry.base_pay = 0  # Only sleepover allowance
        
        # Calculate NDIS charges for sleepover
        # For sleepover shifts, NDIS charges are per-shift based
        roster_entry = calculate_ndis_charges(roster_entry, settings, "sleepover_default")
            
    else:
        # Regular shift calculation
        roster_entry.sleepover_allowance = 0
        
        # Use manual hourly rate if provided
        if roster_entry.manual_hourly_rate:
            hourly_rate = roster_entry.manual_hourly_rate
        else:
            # Use manual shift type if provided, otherwise determine automatically
            if roster_entry.manual_shift_type:
                shift_type_map = {
                    "weekday_day": ShiftType.WEEKDAY_DAY,
                    "weekday_evening": ShiftType.WEEKDAY_EVENING,
                    "weekday_night": ShiftType.WEEKDAY_NIGHT,
                    "saturday": ShiftType.SATURDAY,
                    "sunday": ShiftType.SUNDAY,
                    "public_holiday": ShiftType.PUBLIC_HOLIDAY
                }
                shift_type = shift_type_map.get(roster_entry.manual_shift_type, ShiftType.WEEKDAY_DAY)
            else:
                # Determine shift type automatically
                shift_type = determine_shift_type(
                    roster_entry.date, 
                    roster_entry.start_time, 
                    roster_entry.end_time,
                    roster_entry.is_public_holiday
                )
            
            hourly_rate = get_hourly_rate_for_shift_type(shift_type, settings)
        
        roster_entry.base_pay = hours * hourly_rate
    
    # Calculate NDIS charges - determine shift type for NDIS calculation
    # Skip NDIS calculation for sleepover shifts as it's already calculated above
    is_sleepover = roster_entry.manual_sleepover if roster_entry.manual_sleepover is not None else roster_entry.is_sleepover
    
    if not is_sleepover:
        if roster_entry.manual_shift_type:
            ndis_shift_type = roster_entry.manual_shift_type
        else:
            # Determine shift type automatically for NDIS calculation
            shift_type_enum = determine_shift_type(
                roster_entry.date, 
                roster_entry.start_time, 
                roster_entry.end_time,
                roster_entry.is_public_holiday
            )
            # Convert enum to string for NDIS calculation
            if shift_type_enum == ShiftType.PUBLIC_HOLIDAY:
                ndis_shift_type = "public_holiday"
            elif shift_type_enum == ShiftType.SATURDAY:
                ndis_shift_type = "saturday"
            elif shift_type_enum == ShiftType.SUNDAY:
                ndis_shift_type = "sunday"
            elif shift_type_enum == ShiftType.WEEKDAY_EVENING:
                ndis_shift_type = "weekday_evening"
            elif shift_type_enum == ShiftType.WEEKDAY_NIGHT:
                ndis_shift_type = "weekday_night"
            else:
                ndis_shift_type = "weekday_day"
        
        # Calculate NDIS charges for regular shifts only
        roster_entry = calculate_ndis_charges(roster_entry, settings, ndis_shift_type)
    
    roster_entry.total_pay = roster_entry.base_pay + roster_entry.sleepover_allowance
    return roster_entry

def calculate_pay(roster_entry: RosterEntry, settings: Settings) -> RosterEntry:
    """Calculate pay for a roster entry with cross-midnight logic"""
    return calculate_cross_midnight_pay(roster_entry, settings)

# Initialize default data
def initialize_default_data():
    """Initialize default staff and shift templates"""
    
    # Default staff members
    default_staff = [
        "Angela", "Chanelle", "Rose", "Caroline", "Nox", "Elina",
        "Kayla", "Rhet", "Nikita", "Molly", "Felicity", "Issey"
    ]
    
    for staff_name in default_staff:
        existing = db.staff.find_one({"name": staff_name})
        if not existing:
            staff = Staff(
                id=str(uuid.uuid4()),
                name=staff_name,
                active=True,
                created_at=datetime.now()
            )
            db.staff.insert_one(staff.dict())
    
    # Clear existing shift templates and create new ones per user requirements
    db.shift_templates.delete_many({})
    
    # Updated shift templates according to user specifications
    shift_templates = [
        # Monday
        {"name": "Monday Shift 1", "start_time": "07:30", "end_time": "15:30", "is_sleepover": False, "day_of_week": 0},  # Weekday Day
        {"name": "Monday Shift 2", "start_time": "15:00", "end_time": "20:00", "is_sleepover": False, "day_of_week": 0},  # Weekday Day  
        {"name": "Monday Shift 3", "start_time": "15:30", "end_time": "23:30", "is_sleepover": False, "day_of_week": 0},  # Weekday Evening
        {"name": "Monday Shift 4", "start_time": "23:30", "end_time": "07:30", "is_sleepover": True, "day_of_week": 0},   # Sleepover
        
        # Tuesday
        {"name": "Tuesday Shift 1", "start_time": "07:30", "end_time": "15:30", "is_sleepover": False, "day_of_week": 1},  # Weekday Day
        {"name": "Tuesday Shift 2", "start_time": "12:00", "end_time": "20:00", "is_sleepover": False, "day_of_week": 1},  # Weekday Day
        {"name": "Tuesday Shift 3", "start_time": "15:30", "end_time": "23:30", "is_sleepover": False, "day_of_week": 1},  # Weekday Evening
        {"name": "Tuesday Shift 4", "start_time": "23:30", "end_time": "07:30", "is_sleepover": True, "day_of_week": 1},   # Sleepover
        
        # Wednesday
        {"name": "Wednesday Shift 1", "start_time": "07:30", "end_time": "15:30", "is_sleepover": False, "day_of_week": 2},  # Weekday Day
        {"name": "Wednesday Shift 2", "start_time": "15:00", "end_time": "20:00", "is_sleepover": False, "day_of_week": 2},  # Weekday Day
        {"name": "Wednesday Shift 3", "start_time": "15:30", "end_time": "23:30", "is_sleepover": False, "day_of_week": 2},  # Weekday Evening
        {"name": "Wednesday Shift 4", "start_time": "23:30", "end_time": "07:30", "is_sleepover": True, "day_of_week": 2},   # Sleepover
        
        # Thursday
        {"name": "Thursday Shift 1", "start_time": "07:30", "end_time": "15:30", "is_sleepover": False, "day_of_week": 3},  # Weekday Day
        {"name": "Thursday Shift 2", "start_time": "12:00", "end_time": "20:00", "is_sleepover": False, "day_of_week": 3},  # Weekday Day
        {"name": "Thursday Shift 3", "start_time": "15:30", "end_time": "23:30", "is_sleepover": False, "day_of_week": 3},  # Weekday Evening
        {"name": "Thursday Shift 4", "start_time": "23:30", "end_time": "07:30", "is_sleepover": True, "day_of_week": 3},   # Sleepover
        
        # Friday
        {"name": "Friday Shift 1", "start_time": "07:30", "end_time": "15:30", "is_sleepover": False, "day_of_week": 4},  # Weekday Day
        {"name": "Friday Shift 2", "start_time": "15:00", "end_time": "20:00", "is_sleepover": False, "day_of_week": 4},  # Weekday Day
        {"name": "Friday Shift 3", "start_time": "15:30", "end_time": "23:30", "is_sleepover": False, "day_of_week": 4},  # Weekday Evening
        {"name": "Friday Shift 4", "start_time": "23:30", "end_time": "07:30", "is_sleepover": True, "day_of_week": 4},   # Sleepover
        
        # Saturday
        {"name": "Saturday Shift 1", "start_time": "07:30", "end_time": "15:30", "is_sleepover": False, "day_of_week": 5},  # Saturday Rate
        {"name": "Saturday Shift 2", "start_time": "15:00", "end_time": "20:00", "is_sleepover": False, "day_of_week": 5},  # Saturday Rate
        {"name": "Saturday Shift 3", "start_time": "15:30", "end_time": "23:30", "is_sleepover": False, "day_of_week": 5},  # Saturday Rate
        {"name": "Saturday Shift 4", "start_time": "23:30", "end_time": "07:30", "is_sleepover": True, "day_of_week": 5},   # Sleepover
        
        # Sunday
        {"name": "Sunday Shift 1", "start_time": "07:30", "end_time": "15:30", "is_sleepover": False, "day_of_week": 6},  # Sunday Rate
        {"name": "Sunday Shift 2", "start_time": "15:00", "end_time": "20:00", "is_sleepover": False, "day_of_week": 6},  # Sunday Rate
        {"name": "Sunday Shift 3", "start_time": "15:30", "end_time": "23:30", "is_sleepover": False, "day_of_week": 6},  # Sunday Rate
        {"name": "Sunday Shift 4", "start_time": "23:30", "end_time": "07:30", "is_sleepover": True, "day_of_week": 6},   # Sleepover
    ]
    
    for template_data in shift_templates:
        template = ShiftTemplate(
            id=str(uuid.uuid4()),
            **template_data
        )
        db.shift_templates.insert_one(template.dict())
    
    # Initialize default settings
    existing_settings = db.settings.find_one()
    if not existing_settings:
        settings = Settings()
        db.settings.insert_one(settings.dict())

# Authentication dependency
def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())):
    """Get current authenticated user"""
    token = credentials.credentials
    session = db.sessions.find_one({"token": token, "is_active": True})
    
    if not session or session["expires_at"] < datetime.utcnow():
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    user = db.users.find_one({"id": session["user_id"], "is_active": True})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    
    return user

# API Endpoints

@app.on_event("startup")
async def startup_event():
    initialize_default_data()

@app.get("/api/health")
async def health_check():
    return {"status": "healthy"}

# Staff endpoints
@app.get("/api/staff")
async def get_staff():
    staff_list = list(db.staff.find({"active": True}, {"_id": 0}))
    # Sort staff alphabetically by name
    staff_list.sort(key=lambda staff: staff['name'].lower())
    return staff_list

@app.post("/api/staff")
async def create_staff(staff: Staff):
    """Create a new staff member"""
    if not staff.id:
        staff.id = str(uuid.uuid4())
    if not staff.created_at:
        staff.created_at = datetime.now()
    
    # Validate staff name is not empty
    if not staff.name or staff.name.strip() == "":
        raise HTTPException(status_code=422, detail="Staff name cannot be empty")
    
    # Check if staff name already exists
    existing_staff = db.staff.find_one({"name": staff.name, "active": True})
    if existing_staff:
        raise HTTPException(status_code=400, detail=f"Staff member with name '{staff.name}' already exists")
    
    try:
        db.staff.insert_one(staff.dict())
        return staff
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create staff member: {str(e)}")

@app.put("/api/staff/{staff_id}")
async def update_staff(staff_id: str, staff: Staff):
    result = db.staff.update_one({"id": staff_id}, {"$set": staff.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Staff not found")
    return staff

@app.delete("/api/staff/{staff_id}")
async def delete_staff(staff_id: str, current_user: dict = Depends(get_current_user)):
    """Deactivate a staff member (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if staff member exists
    staff_member = db.staff.find_one({"id": staff_id})
    if not staff_member:
        raise HTTPException(status_code=404, detail="Staff not found")
    
    # Check for assigned shifts
    assigned_shifts = list(db.roster.find({"staff_id": staff_id}))
    future_shifts = []
    past_shifts = []
    
    from datetime import datetime
    today = datetime.now().strftime("%Y-%m-%d")
    
    for shift in assigned_shifts:
        if shift.get("date", "") >= today:
            future_shifts.append(shift)
        else:
            past_shifts.append(shift)
    
    # Deactivate the staff member
    result = db.staff.update_one({"id": staff_id}, {"$set": {"active": False}})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Staff not found")
    
    # Unassign from future shifts (keep past shifts for record keeping)
    if future_shifts:
        db.roster.update_many(
            {"staff_id": staff_id, "date": {"$gte": today}},
            {"$unset": {"staff_id": "", "staff_name": ""}}
        )
    
    response = {
        "message": f"Staff member '{staff_member.get('first_name', '')} {staff_member.get('last_name', '')}' has been deactivated",
        "staff_name": f"{staff_member.get('first_name', '')} {staff_member.get('last_name', '')}".strip(),
        "shifts_affected": {
            "future_shifts_unassigned": len(future_shifts),
            "past_shifts_preserved": len(past_shifts),
            "total_shifts": len(assigned_shifts)
        }
    }
    
    return response

# Shift template endpoints
@app.get("/api/shift-templates")
async def get_shift_templates():
    """Get shift templates with calculated rates and shift types"""
    templates = list(db.shift_templates.find({}, {"_id": 0}))
    
    # Get current settings for rate calculations
    settings = db.settings.find_one() or {}
    rates = settings.get("rates", {})
    
    # Enhance each template with calculated information
    enhanced_templates = []
    for template in templates:
        # Create a sample date for rate calculation (use today's date adjusted for day_of_week)
        today = datetime.now()
        # Calculate date for the template's day of week
        days_ahead = template.get('day_of_week', 0) - today.weekday()
        if days_ahead <= 0:  # Target day already happened this week
            days_ahead += 7
        target_date = today + timedelta(days=days_ahead)
        date_str = target_date.strftime("%Y-%m-%d")
        
        # Calculate shift type using the proper backend logic
        shift_type = determine_shift_type(
            date_str,
            template.get('start_time', '00:00'),
            template.get('end_time', '00:00'),
            False  # Not a public holiday for template calculation
        )
        
        # Get the hourly rate for this shift type
        rate_map = {
            ShiftType.WEEKDAY_DAY: rates.get('weekday_day', 42.00),
            ShiftType.WEEKDAY_EVENING: rates.get('weekday_evening', 44.50),
            ShiftType.WEEKDAY_NIGHT: rates.get('weekday_night', 48.50),
            ShiftType.SATURDAY: rates.get('saturday', 57.50),
            ShiftType.SUNDAY: rates.get('sunday', 74.00),
            ShiftType.PUBLIC_HOLIDAY: rates.get('public_holiday', 88.50),
            ShiftType.SLEEPOVER: rates.get('sleepover_default', 175.00)
        }
        
        hourly_rate = rate_map.get(shift_type, 42.00)
        
        # Calculate hours and total pay
        hours_worked = calculate_hours_worked(
            template.get('start_time', '00:00'),
            template.get('end_time', '00:00')
        )
        
        # Handle sleepover shifts differently
        if template.get('is_sleepover', False):
            shift_type = ShiftType.SLEEPOVER
            hourly_rate = rates.get('sleepover_default', 175.00)
            total_pay = hourly_rate  # Sleepover is flat rate
            rate_display = "Sleepover"
        else:
            total_pay = hours_worked * hourly_rate
            rate_display = f"${hourly_rate:.2f}/hr"
        
        # Add calculated fields to template
        enhanced_template = dict(template)
        enhanced_template.update({
            'calculated_shift_type': shift_type.value if hasattr(shift_type, 'value') else str(shift_type),
            'calculated_hourly_rate': hourly_rate,
            'calculated_hours': hours_worked,
            'calculated_total_pay': total_pay,
            'rate_display': rate_display,
            'shift_type_display': shift_type.value if hasattr(shift_type, 'value') else str(shift_type)
        })
        
        enhanced_templates.append(enhanced_template)
    
    return enhanced_templates

@app.post("/api/shift-templates")
async def create_shift_template(template: ShiftTemplate):
    template.id = str(uuid.uuid4())
    db.shift_templates.insert_one(template.dict())
    return template

@app.put("/api/shift-templates/{template_id}")
async def update_shift_template(template_id: str, template: ShiftTemplate):
    """Update a shift template with all fields including manual overrides"""
    result = db.shift_templates.update_one({"id": template_id}, {"$set": template.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Shift template not found")
    return template

# Calendar events endpoints
@app.get("/api/calendar-events")
async def get_calendar_events(start_date: Optional[str] = None, end_date: Optional[str] = None, event_type: Optional[str] = None):
    """Get calendar events with optional filtering"""
    query = {"is_active": True}
    
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    
    if event_type:
        query["event_type"] = event_type
    
    events = list(db.calendar_events.find(query, {"_id": 0}))
    return events

@app.get("/api/calendar-events/{date}")
async def get_events_for_date(date: str):
    """Get all events for a specific date"""
    events = list(db.calendar_events.find({"date": date, "is_active": True}, {"_id": 0}))
    return events

@app.post("/api/calendar-events")
async def create_calendar_event(event: CalendarEvent):
    """Create a new calendar event"""
    event.id = str(uuid.uuid4())
    event.created_at = datetime.now()
    db.calendar_events.insert_one(event.dict())
    return event

@app.put("/api/calendar-events/{event_id}")
async def update_calendar_event(event_id: str, event: CalendarEvent):
    """Update an existing calendar event"""
    result = db.calendar_events.update_one({"id": event_id}, {"$set": event.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Calendar event not found")
    return event

@app.delete("/api/calendar-events/{event_id}")
async def delete_calendar_event(event_id: str):
    """Delete a calendar event"""
    result = db.calendar_events.update_one({"id": event_id}, {"$set": {"is_active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Calendar event not found")
    return {"message": "Calendar event deleted"}

@app.put("/api/calendar-events/{event_id}/complete")
async def complete_task(event_id: str):
    """Mark a task as completed"""
    result = db.calendar_events.update_one(
        {"id": event_id, "event_type": "task"}, 
        {"$set": {"is_completed": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task marked as completed"}

# Day template endpoints
@app.get("/api/day-templates")
async def get_day_templates():
    """Get all day templates"""
    templates = list(db.day_templates.find({"is_active": True}, {"_id": 0}))
    return templates

@app.get("/api/day-templates/{day_of_week}")
async def get_day_templates_for_day(day_of_week: int):
    """Get day templates for a specific day of week"""
    templates = list(db.day_templates.find({"day_of_week": day_of_week, "is_active": True}, {"_id": 0}))
    return templates

@app.post("/api/day-templates")
async def create_day_template(template: DayTemplate):
    """Create a new day template"""
    template.id = str(uuid.uuid4())
    template.created_at = datetime.now()
    db.day_templates.insert_one(template.dict())
    return template

@app.post("/api/day-templates/save-day/{template_name}")
async def save_day_as_template(template_name: str, date: str):
    """Save all shifts from a specific date as a day template"""
    # Get all roster entries for the specific date
    roster_entries = list(db.roster.find({"date": date}, {"_id": 0}))
    
    if not roster_entries:
        raise HTTPException(status_code=404, detail=f"No shifts found for date {date}")
    
    # Get day of week from the date
    date_obj = datetime.strptime(date, "%Y-%m-%d")
    day_of_week = date_obj.weekday()  # 0=Monday, 6=Sunday
    day_name = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'][day_of_week]
    
    # Extract shift data (without staff assignments and date-specific info)
    shifts = []
    for entry in roster_entries:
        shift_data = {
            "start_time": entry["start_time"],
            "end_time": entry["end_time"],
            "is_sleepover": entry.get("is_sleepover", False)
        }
        shifts.append(shift_data)
    
    # Sort shifts by start time for consistency
    shifts.sort(key=lambda x: x["start_time"])
    
    # Create the day template
    day_template = DayTemplate(
        id=str(uuid.uuid4()),
        name=template_name,
        description=f"{day_name} template with {len(shifts)} shifts (from {date})",
        day_of_week=day_of_week,
        shifts=shifts,
        created_at=datetime.now()
    )
    
    db.day_templates.insert_one(day_template.dict())
    return day_template

@app.post("/api/day-templates/apply-to-date/{template_id}")
async def apply_day_template_to_date(template_id: str, target_date: str):
    """Apply a day template to a specific date"""
    # Get the day template
    template_doc = db.day_templates.find_one({"id": template_id, "is_active": True})
    if not template_doc:
        raise HTTPException(status_code=404, detail="Day template not found")
    
    template = DayTemplate(**template_doc)
    
    # Check if target date already has shifts and look for overlaps
    overlaps = []
    for shift_data in template.shifts:
        if check_shift_overlap(target_date, shift_data["start_time"], shift_data["end_time"], shift_name=template.name):
            overlaps.append(f"{shift_data['start_time']}-{shift_data['end_time']}")
    
    if overlaps:
        raise HTTPException(
            status_code=409, 
            detail=f"Cannot apply template: overlaps detected with existing shifts at times: {', '.join(overlaps)} (Note: 2:1 shifts are allowed to overlap)"
        )
    
    # Apply the template shifts to the target date
    entries_created = 0
    settings_doc = db.settings.find_one()
    settings = Settings(**settings_doc) if settings_doc else Settings()
    
    for shift_data in template.shifts:
        # Create roster entry
        entry = RosterEntry(
            id=str(uuid.uuid4()),
            date=target_date,
            shift_template_id=f"day-template-{template_id}",
            start_time=shift_data["start_time"],
            end_time=shift_data["end_time"],
            is_sleepover=shift_data.get("is_sleepover", False)
        )
        
        # Calculate pay
        entry = calculate_pay(entry, settings)
        
        db.roster.insert_one(entry.dict())
        entries_created += 1
    
    return {
        "message": f"Applied '{template.name}' to {target_date}",
        "entries_created": entries_created,
        "template_name": template.name
    }

@app.delete("/api/day-templates/{template_id}")
async def delete_day_template(template_id: str):
    """Delete a day template"""
    result = db.day_templates.update_one({"id": template_id}, {"$set": {"is_active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Day template not found")
    return {"message": "Day template deleted"}

@app.post("/api/generate-roster-from-shift-templates/{month}")
async def generate_roster_from_shift_templates(month: str, templates_data: dict):
    """Generate roster for a month using the shift templates from Shift Times section"""
    year, month_num = map(int, month.split("-"))
    templates = templates_data.get("templates", [])
    force_overlaps = templates_data.get("force_overlaps", False)
    
    if not templates:
        raise HTTPException(status_code=400, detail="No shift templates provided")
    
    # Generate entries for each day of the month using shift templates
    from calendar import monthrange
    _, days_in_month = monthrange(year, month_num)
    
    entries_created = 0
    overlaps_detected = []
    
    for day in range(1, days_in_month + 1):
        date_obj = datetime(year, month_num, day)
        date_str = date_obj.strftime("%Y-%m-%d")
        day_of_week = date_obj.weekday()  # 0=Monday, 6=Sunday
        
        # Get shift templates for this day of week
        day_templates = [t for t in templates if t.get("day_of_week") == day_of_week]
        
        for template in day_templates:
            template_name = template.get("name", "")
            allow_overlap = template.get("allow_overlap", False)
            
            # Check for overlaps (unless forced or explicitly allowed)
            if not force_overlaps and not allow_overlap:
                if check_shift_overlap(date_str, template["start_time"], template["end_time"], shift_name=template_name):
                    overlaps_detected.append({
                        "date": date_str,
                        "start_time": template["start_time"],
                        "end_time": template["end_time"],
                        "name": template_name,
                        "reason": "Overlap detected in shift templates"
                    })
                    continue  # Skip overlapping shifts unless forced
            
            # Check if entry already exists (unless forcing overlaps)
            existing = db.roster.find_one({
                "date": date_str,
                "start_time": template["start_time"],
                "end_time": template["end_time"]
            }) if not force_overlaps else None
            
            if not existing or force_overlaps:
                entry = RosterEntry(
                    id=str(uuid.uuid4()),
                    date=date_str,
                    shift_template_id=template.get("id", f"template-{day_of_week}"),
                    start_time=template["start_time"],
                    end_time=template["end_time"],
                    is_sleepover=template.get("is_sleepover", False),
                    manual_shift_type=template.get("manual_shift_type"),
                    manual_hourly_rate=template.get("manual_hourly_rate"),
                    allow_overlap=allow_overlap or force_overlaps
                )
                
                # Calculate pay using template overrides
                settings_doc = db.settings.find_one()
                settings = Settings(**settings_doc) if settings_doc else Settings()
                entry = calculate_pay(entry, settings)
                
                db.roster.insert_one(entry.dict())
                entries_created += 1
    
    result = {
        "message": f"Generated {entries_created} roster entries for {month} using Shift Times templates",
        "entries_created": entries_created,
        "force_overlaps": force_overlaps
    }
    
    if overlaps_detected:
        result["overlaps_detected"] = len(overlaps_detected)
        result["overlap_details"] = overlaps_detected[:5]  # Show first 5 overlaps
        if not force_overlaps:
            result["note"] = "Shifts with overlaps were skipped. Use 'Publish All' to force overlaps."
    
    if force_overlaps:
        result["message"] += " (overlaps forced)"
    
    return result

# Roster template endpoints
@app.get("/api/roster-templates")
async def get_roster_templates():
    """Get all roster templates"""
    templates = list(db.roster_templates.find({"is_active": True}, {"_id": 0}))
    return templates

@app.post("/api/roster-templates")
async def create_roster_template(template: RosterTemplate):
    """Create a new roster template"""
    template.id = str(uuid.uuid4())
    template.created_at = datetime.now()
    db.roster_templates.insert_one(template.dict())
    return template

@app.put("/api/roster-templates/{template_id}")
async def update_roster_template(template_id: str, template: RosterTemplate):
    """Update an existing roster template"""
    result = db.roster_templates.update_one({"id": template_id}, {"$set": template.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Roster template not found")
    return template

@app.delete("/api/roster-templates/{template_id}")
async def delete_roster_template(template_id: str):
    """Delete a roster template"""
    result = db.roster_templates.update_one({"id": template_id}, {"$set": {"is_active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Roster template not found")
    return {"message": "Roster template deleted"}

@app.post("/api/roster-templates/save-current/{template_name}")
async def save_current_roster_as_template(template_name: str, month: str):
    """Save current month's roster as a template"""
    # Get all roster entries for the month
    roster_entries = list(db.roster.find({"date": {"$regex": f"^{month}"}}, {"_id": 0}))
    
    if not roster_entries:
        raise HTTPException(status_code=404, detail="No roster entries found for the specified month")
    
    # Group entries by day of week
    template_data = {}
    
    for entry in roster_entries:
        date_obj = datetime.strptime(entry["date"], "%Y-%m-%d")
        day_of_week = str(date_obj.weekday())  # Convert to string for MongoDB compatibility
        
        if day_of_week not in template_data:
            template_data[day_of_week] = []
        
        # Store the shift template info (without staff assignments and dates)
        shift_template = {
            "start_time": entry["start_time"],
            "end_time": entry["end_time"],
            "is_sleepover": entry.get("is_sleepover", False)
        }
        
        # Avoid duplicates by checking if this shift template already exists for this day
        if shift_template not in template_data[day_of_week]:
            template_data[day_of_week].append(shift_template)
    
    # Create the roster template
    roster_template = RosterTemplate(
        id=str(uuid.uuid4()),
        name=template_name,
        description=f"Template created from {month} roster",
        created_at=datetime.now(),
        template_data=template_data
    )
    
    db.roster_templates.insert_one(roster_template.dict())
    return roster_template

@app.post("/api/generate-roster-from-template/{template_id}/{month}")
async def generate_roster_from_template(template_id: str, month: str, force_overlaps: bool = False):
    """Generate roster entries for a month using a roster template with advanced 2:1 shift support"""
    # Get the roster template
    template_doc = db.roster_templates.find_one({"id": template_id, "is_active": True})
    if not template_doc:
        raise HTTPException(status_code=404, detail="Roster template not found")
    
    template = RosterTemplate(**template_doc)
    year, month_num = map(int, month.split("-"))
    
    # Generate entries for each day of the month
    from calendar import monthrange
    _, days_in_month = monthrange(year, month_num)
    
    entries_created = 0
    overlaps_detected = []
    duplicates_prevented = []
    duplicates_allowed = []
    
    for day in range(1, days_in_month + 1):
        date_obj = datetime(year, month_num, day)
        date_str = date_obj.strftime("%Y-%m-%d")
        day_of_week = date_obj.weekday()  # 0=Monday, 6=Sunday
        
        # Get shift templates for this day of week
        day_shifts = template.template_data.get(str(day_of_week), [])
        if not day_shifts:
            day_shifts = template.template_data.get(day_of_week, [])  # Try integer key
        
        for shift_data in day_shifts:
            # Get shift name from template if available
            shift_name = shift_data.get("name", template.name)
            
            # Check for existing entries at same time
            existing_entries = list(db.roster.find({
                "date": date_str,
                "start_time": shift_data["start_time"],
                "end_time": shift_data["end_time"]
            }))
            
            should_skip = False
            skip_reason = ""
            
            # Apply duplicate prevention rules based on template configuration (unless forcing overlaps)
            if existing_entries and not force_overlaps:
                # Check if any existing entries are unassigned
                unassigned_exists = any(not entry.get("staff_id") and not entry.get("staff_name") for entry in existing_entries)
                
                if unassigned_exists and template.prevent_duplicate_unassigned:
                    should_skip = True
                    skip_reason = "Duplicate prevented: Unassigned shift already exists"
                    duplicates_prevented.append({
                        "date": date_str,
                        "start_time": shift_data["start_time"],
                        "end_time": shift_data["end_time"],
                        "reason": skip_reason
                    })
                
                # Check if we should allow different staff only
                elif template.allow_different_staff_only and not template.enable_2_1_shift:
                    # For non-2:1 templates with different staff rule, check if all slots are filled with different staff
                    assigned_staff = [entry.get("staff_id") for entry in existing_entries if entry.get("staff_id")]
                    if len(assigned_staff) >= 1:  # Already has one staff member
                        should_skip = True
                        skip_reason = "Single staff per shift (non-2:1 template)"
                        duplicates_prevented.append({
                            "date": date_str,
                            "start_time": shift_data["start_time"],
                            "end_time": shift_data["end_time"],
                            "reason": skip_reason
                        })
            
            if should_skip:
                continue
            
            # Check for overlaps if not overridden and not forcing overlaps
            if not template.allow_overlap_override and not force_overlaps:
                overlap_check = check_shift_overlap(date_str, shift_data["start_time"], shift_data["end_time"], shift_name=shift_name)
                if overlap_check and not template.enable_2_1_shift:
                    overlaps_detected.append({
                        "date": date_str,
                        "start_time": shift_data["start_time"],
                        "end_time": shift_data["end_time"],
                        "name": shift_name,
                        "reason": "Overlap detected and not overridden"
                    })
                    if not force_overlaps:
                        continue  # Skip overlapping shifts unless forced
            
            # Create new entry
            entry = RosterEntry(
                id=str(uuid.uuid4()),
                date=date_str,
                shift_template_id=f"template-{template_id}-{day_of_week}",
                start_time=shift_data["start_time"],
                end_time=shift_data["end_time"],
                is_sleepover=shift_data.get("is_sleepover", False),
                allow_overlap=template.enable_2_1_shift or template.allow_overlap_override or force_overlaps
            )
            
            # Calculate pay
            settings_doc = db.settings.find_one()
            settings = Settings(**settings_doc) if settings_doc else Settings()
            entry = calculate_pay(entry, settings)
            
            db.roster.insert_one(entry.dict())
            entries_created += 1
            
            # Track allowed duplicates for reporting
            if existing_entries:
                duplicates_allowed.append({
                    "date": date_str,
                    "start_time": shift_data["start_time"],
                    "end_time": shift_data["end_time"],
                    "reason": "2:1 shift, different staff allowed, or forced overlaps"
                })
    
    result = {
        "message": f"Generated {entries_created} roster entries for {month} using template '{template.name}'",
        "entries_created": entries_created,
        "force_overlaps": force_overlaps,
        "template_config": {
            "enable_2_1_shift": template.enable_2_1_shift,
            "allow_overlap_override": template.allow_overlap_override,
            "prevent_duplicate_unassigned": template.prevent_duplicate_unassigned,
            "allow_different_staff_only": template.allow_different_staff_only
        }
    }
    
    if overlaps_detected:
        result["overlaps_detected"] = len(overlaps_detected)
        result["overlap_details"] = overlaps_detected[:5]
    
    if duplicates_prevented:
        result["duplicates_prevented"] = len(duplicates_prevented)
        result["prevention_details"] = duplicates_prevented[:5]
    
    if duplicates_allowed:
        result["duplicates_allowed"] = len(duplicates_allowed)
        result["allowance_details"] = duplicates_allowed[:5]
    
    if force_overlaps:
        result["message"] += " (overlaps forced)"
    
    return result

# Roster endpoints
@app.get("/api/roster")
async def get_roster(month: str, current_user: dict = Depends(get_current_user)):
    """Get roster for a specific month (YYYY-MM format) with role-based filtering and pay privacy"""
    
    # Base query for the specified month
    query = {"date": {"$regex": f"^{month}"}}
    
    # Get all roster entries for the month (no filtering by staff for staff users anymore)
    roster_entries = list(db.roster.find(query, {"_id": 0}))
    
    # Apply pay filtering for staff users
    if current_user["role"] == "staff":
        staff_id = current_user.get("staff_id") or current_user.get("id")
        
        # Filter pay information based on staff access rules:
        # - Staff can see pay for their own shifts
        # - Staff can see pay for unassigned shifts 
        # - Staff cannot see pay for other staff's shifts
        for entry in roster_entries:
            # If shift is assigned to someone else (not this staff member and not unassigned)
            if entry.get("staff_id") and entry.get("staff_id") != staff_id:
                # Remove pay information for other staff's shifts
                entry["total_pay"] = None
                entry["base_pay"] = None
                entry["sleepover_allowance"] = None
                entry["ndis_total_charge"] = None
                entry["ndis_hourly_charge"] = None
                entry["ndis_shift_charge"] = None
                # Keep staff_name, hours_worked, time info for display
            # For own shifts and unassigned shifts, keep all pay information intact
    
    return roster_entries

@app.post("/api/roster")
async def create_roster_entry(entry: RosterEntry):
    # Get current settings for pay calculation
    settings_doc = db.settings.find_one()
    settings = Settings(**settings_doc) if settings_doc else Settings()
    
    entry.id = str(uuid.uuid4())
    entry = calculate_pay(entry, settings)
    
    db.roster.insert_one(entry.dict())
    return entry

@app.put("/api/roster/{entry_id}")
async def update_roster_entry(entry_id: str, entry: RosterEntry):
    # Get shift name from template if available
    shift_name = ""
    if entry.shift_template_id:
        template = db.shift_templates.find_one({"id": entry.shift_template_id})
        if template:
            shift_name = template.get("name", "")
    
    # Check for overlaps (excluding current entry, allows 2:1 shifts and manual override)
    if not entry.allow_overlap and check_shift_overlap(entry.date, entry.start_time, entry.end_time, exclude_id=entry_id, shift_name=shift_name):
        raise HTTPException(
            status_code=409, 
            detail=f"Updated shift would overlap with existing shift on {entry.date}"
        )
    
    # Get current settings for pay calculation
    settings_doc = db.settings.find_one()
    settings = Settings(**settings_doc) if settings_doc else Settings()
    
    entry = calculate_pay(entry, settings)
    
    result = db.roster.update_one({"id": entry_id}, {"$set": entry.dict()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Roster entry not found")
    return entry

@app.delete("/api/roster/{entry_id}")
async def delete_roster_entry(entry_id: str):
    result = db.roster.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Roster entry not found")
    return {"message": "Roster entry deleted"}

@app.post("/api/admin/migrate-ndis-charges")
async def migrate_ndis_charges_to_existing_entries(current_user: dict = Depends(get_current_user)):
    """Migrate NDIS charge calculations to existing roster entries"""
    # Admin only endpoint
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get current settings
    settings_doc = db.settings.find_one()
    settings = Settings(**settings_doc) if settings_doc else Settings()
    
    # Find all roster entries that don't have NDIS fields or have zero values
    roster_entries = list(db.roster.find())
    updated_count = 0
    errors = []
    
    for entry_doc in roster_entries:
        try:
            # Check if entry needs NDIS migration
            needs_migration = (
                "ndis_total_charge" not in entry_doc or 
                entry_doc.get("ndis_total_charge", 0) == 0 or
                not entry_doc.get("ndis_line_item_code")
            )
            
            if needs_migration:
                # Convert to RosterEntry object
                roster_entry = RosterEntry(**entry_doc)
                
                # Recalculate pay (including NDIS charges)
                roster_entry = calculate_pay(roster_entry, settings)
                
                # Update in database
                db.roster.update_one(
                    {"id": roster_entry.id}, 
                    {"$set": roster_entry.dict()}
                )
                updated_count += 1
                
        except Exception as e:
            errors.append(f"Entry {entry_doc.get('id', 'unknown')}: {str(e)}")
    
    return {
        "message": f"NDIS charge migration completed",
        "entries_updated": updated_count,
        "total_entries": len(roster_entries),
        "errors": errors
    }

# Settings endpoints
@app.get("/api/settings")
async def get_settings():
    settings_doc = db.settings.find_one({}, {"_id": 0})
    return settings_doc if settings_doc else Settings().dict()

@app.put("/api/settings")
async def update_settings(settings: Settings):
    db.settings.update_one({}, {"$set": settings.dict()}, upsert=True)
    return settings

# Generate monthly roster
@app.post("/api/generate-roster/{month}")
async def generate_monthly_roster(month: str):
    """Generate roster entries for a month based on shift templates"""
    year, month_num = map(int, month.split("-"))
    
    # Get shift templates
    templates = list(db.shift_templates.find())
    
    # Generate entries for each day of the month
    from calendar import monthrange
    _, days_in_month = monthrange(year, month_num)
    
    entries_created = 0
    for day in range(1, days_in_month + 1):
        date_obj = datetime(year, month_num, day)
        date_str = date_obj.strftime("%Y-%m-%d")
        day_of_week = date_obj.weekday()  # 0=Monday
        
        # Find templates for this day of week
        day_templates = [t for t in templates if t["day_of_week"] == day_of_week]
        
        for template in day_templates:
            # Check if entry already exists
            existing = db.roster.find_one({
                "date": date_str,
                "shift_template_id": template["id"]
            })
            
            if not existing:
                entry = RosterEntry(
                    id=str(uuid.uuid4()),
                    date=date_str,
                    shift_template_id=template["id"],
                    start_time=template["start_time"],
                    end_time=template["end_time"],
                    is_sleepover=template["is_sleepover"]
                )
                
                # Calculate pay
                settings_doc = db.settings.find_one()
                settings = Settings(**settings_doc) if settings_doc else Settings()
                entry = calculate_pay(entry, settings)
                
                db.roster.insert_one(entry.dict())
                entries_created += 1
    
    return {"message": f"Generated {entries_created} roster entries for {month}"}

# Clear roster for a month
@app.delete("/api/roster/month/{month}")
async def clear_monthly_roster(month: str):
    """Clear all roster entries for a specific month"""
    result = db.roster.delete_many({"date": {"$regex": f"^{month}"}})
    return {"message": f"Deleted {result.deleted_count} roster entries for {month}"}

@app.post("/api/roster/add-shift")
async def add_individual_shift(entry: RosterEntry):
    """Add a single shift to the roster with overlap detection (allows 2:1 shifts and manual override)"""
    # Get shift name from template if available
    shift_name = ""
    if entry.shift_template_id:
        template = db.shift_templates.find_one({"id": entry.shift_template_id})
        if template:
            shift_name = template.get("name", "")
    
    # Check for overlaps (allows 2:1 shifts to overlap, or if allow_overlap is explicitly set)
    if not entry.allow_overlap and check_shift_overlap(entry.date, entry.start_time, entry.end_time, shift_name=shift_name):
        raise HTTPException(
            status_code=409, 
            detail=f"Shift overlaps with existing shift on {entry.date} from {entry.start_time} to {entry.end_time}. Use 'Allow Overlap' option for 2:1 shifts."
        )
    
    # Get current settings for pay calculation
    settings_doc = db.settings.find_one()
    settings = Settings(**settings_doc) if settings_doc else Settings()
    
    entry.id = str(uuid.uuid4())
    entry = calculate_pay(entry, settings)
    
    db.roster.insert_one(entry.dict())
    return entry

# Authentication endpoints
@app.post("/api/auth/login")
async def login(request: LoginRequest):
    """Authenticate user with username and PIN"""
    user = db.users.find_one({"username": request.username, "is_active": True})
    if not user or not verify_pin(request.pin, user["pin_hash"]):
        raise HTTPException(status_code=401, detail="Invalid username or PIN")
    
    # Create session
    token = generate_token()
    session = Session(
        id=str(uuid.uuid4()),
        user_id=user["id"],
        token=token,
        created_at=datetime.utcnow(),
        expires_at=datetime.utcnow() + timedelta(hours=8)  # 8-hour session
    )
    
    db.sessions.insert_one(session.dict())
    
    # Update last login
    db.users.update_one(
        {"id": user["id"]},
        {"$set": {"last_login": datetime.utcnow()}}
    )
    
    # Remove sensitive data from response
    user_data = {k: v for k, v in user.items() if k not in ["pin_hash", "_id"]}
    
    return {
        "user": user_data,
        "token": token,
        "expires_at": session.expires_at
    }

@app.post("/api/auth/change-pin")
async def change_pin(request: ChangePinRequest, user: dict = Depends(get_current_user)):
    """Change user PIN"""
    if not verify_pin(request.current_pin, user["pin_hash"]):
        raise HTTPException(status_code=400, detail="Current PIN is incorrect")
    
    # Validate new PIN (4 or 6 digits)
    if not request.new_pin.isdigit() or len(request.new_pin) not in [4, 6]:
        raise HTTPException(status_code=400, detail="PIN must be 4 or 6 digits")
    
    new_pin_hash = hash_pin(request.new_pin)
    db.users.update_one(
        {"id": user["id"]},
        {"$set": {"pin_hash": new_pin_hash, "is_first_login": False}}
    )
    
    return {"message": "PIN changed successfully"}

@app.put("/api/auth/change-pin")
async def change_user_pin(new_pin_data: dict, current_user: dict = Depends(get_current_user)):
    """Change user PIN (enhanced version for new login system)"""
    new_pin = new_pin_data.get("new_pin")
    
    if not new_pin:
        raise HTTPException(status_code=400, detail="New PIN is required")
    
    # Validate new PIN (4 or 6 digits)
    if not new_pin.isdigit() or len(new_pin) not in [4, 6]:
        raise HTTPException(status_code=400, detail="PIN must be 4 or 6 digits")
    
    # Hash the new PIN
    new_pin_hash = hash_pin(new_pin)
    
    # Update user PIN and mark as not first-time login
    result = db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {
            "pin": new_pin,  # Store plain PIN for display (in production, consider removing)
            "pin_hash": new_pin_hash,
            "is_first_login": False,
            "updated_at": datetime.utcnow()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"success": True, "message": "PIN changed successfully"}

@app.put("/api/auth/reset-pin")
async def admin_reset_user_pin(reset_data: dict, current_user: dict = Depends(get_current_user)):
    """Admin reset user PIN to default (Admin only)"""
    
    # Check if current user is admin
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only administrators can reset PINs")
    
    target_user_id = reset_data.get("user_id")
    if not target_user_id:
        raise HTTPException(status_code=400, detail="User ID is required")
    
    # Find the target user
    target_user = db.users.find_one({"id": target_user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="Target user not found")
    
    # Determine default PIN based on role
    if target_user.get("role") == "admin":
        default_pin = "1234"
    else:  # staff or other roles
        default_pin = "888888"
    
    # Hash the default PIN
    default_pin_hash = hash_pin(default_pin)
    
    # Reset user PIN to default
    result = db.users.update_one(
        {"id": target_user_id},
        {"$set": {
            "pin": default_pin,
            "pin_hash": default_pin_hash,
            "is_first_login": True,  # Mark as first-time login to force PIN change
            "updated_at": datetime.utcnow()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Failed to reset PIN")
    
    return {
        "success": True, 
        "message": f"PIN reset to default ({default_pin}) successfully",
        "default_pin": default_pin
    }

@app.post("/api/auth/reset-pin")
async def reset_pin(request: ResetPinRequest):
    """Request PIN reset via email"""
    user = db.users.find_one({"username": request.username, "email": request.email, "is_active": True})
    if not user:
        raise HTTPException(status_code=404, detail="User not found with provided username and email")
    
    # Generate temporary PIN
    temp_pin = str(secrets.randbelow(1000000)).zfill(6)  # 6-digit temp PIN
    temp_pin_hash = hash_pin(temp_pin)
    
    # Update user with temporary PIN
    db.users.update_one(
        {"id": user["id"]},
        {"$set": {"pin_hash": temp_pin_hash, "is_first_login": True}}
    )
    
    # Send reset email
    send_reset_email(request.email, temp_pin)
    
    return {"message": "Temporary PIN sent to email address"}

@app.post("/api/admin/reset_pin")
async def admin_reset_pin(request: dict, current_user: dict = Depends(get_current_user)):
    """Admin reset PIN for any user (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    email = request.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    
    # Try to find user by email first, then by generated email pattern
    user = db.users.find_one({"email": email, "is_active": True})
    
    # If not found and email looks like generated email, try to find by staff name
    if not user and "@company.com" in email:
        # Extract staff name from generated email pattern
        staff_name_pattern = email.split("@")[0]
        # Find staff member by name pattern
        staff_member = db.staff.find_one({"active": True})
        if staff_member:
            staff_names = list(db.staff.find({"active": True}))
            for staff in staff_names:
                generated_email = f"{staff['name'].lower().replace(' ', '')}@company.com"
                if generated_email == email:
                    # Create a user account for this staff member if it doesn't exist
                    existing_user = db.users.find_one({"staff_id": staff["id"]}) 
                    if not existing_user:
                        # Create user account for staff member with default staff PIN
                        new_user = User(
                            id=str(uuid.uuid4()),
                            username=staff["name"].lower().replace(" ", ""),
                            pin_hash=hash_pin("888888"),  # Default staff PIN: 888888
                            role=UserRole.STAFF,
                            email=email,
                            first_name=staff["name"].split()[0] if " " in staff["name"] else staff["name"],
                            last_name=" ".join(staff["name"].split()[1:]) if " " in staff["name"] else "",
                            staff_id=staff["id"],
                            created_at=datetime.utcnow(),
                            is_first_login=True  # Staff must change PIN on first login
                        )
                        db.users.insert_one(new_user.dict())
                        user = new_user.dict()
                    else:
                        user = existing_user
                    break
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Determine appropriate reset PIN based on user role
    if user.get("role") == "admin":
        reset_pin = "0000"  # Admin reset PIN
        pin_length = 4
    else:
        reset_pin = "888888"  # Staff reset PIN
        pin_length = 6
    
    reset_pin_hash = hash_pin(reset_pin)
    
    # Update user with reset PIN and mark as first login for staff
    update_data = {
        "pin_hash": reset_pin_hash
    }
    
    # Staff must change PIN after reset, Admin doesn't need to
    if user.get("role") != "admin":
        update_data["is_first_login"] = True
    
    db.users.update_one(
        {"id": user["id"]},
        {"$set": update_data}
    )
    
    return {
        "message": "PIN reset successful",
        "temp_pin": reset_pin,
        "username": user.get("username", ""),
        "pin_length": pin_length,
        "must_change": user.get("role") != "admin"
    }

@app.get("/api/auth/logout")
async def logout(token: str):
    """Logout user and invalidate session"""
    db.sessions.update_one(
        {"token": token},
        {"$set": {"is_active": False}}
    )
    return {"message": "Logged out successfully"}

# User management endpoints
@app.get("/api/users/me")
async def get_current_user_profile(current_user: dict = Depends(get_current_user)):
    """Get current user's profile"""
    user_data = {k: v for k, v in current_user.items() if k not in ["pin_hash", "_id"]}
    return user_data

@app.put("/api/users/me")
async def update_current_user_profile(profile_data: dict, current_user: dict = Depends(get_current_user)):
    """Update current user's profile"""
    # Filter allowed fields that user can update
    allowed_fields = ["first_name", "last_name", "email", "phone", "address", "date_of_birth", "emergency_contact"]
    update_data = {k: v for k, v in profile_data.items() if k in allowed_fields}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    # Update user in database
    result = db.users.update_one(
        {"id": current_user["id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Return updated user data
    updated_user = db.users.find_one({"id": current_user["id"]}, {"_id": 0, "pin_hash": 0})
    return updated_user

@app.get("/api/users/login")
async def get_login_users():
    """Get available users for login dropdown (public endpoint)"""
    try:
        # Get only active users with basic info needed for login
        users = list(db.users.find(
            {"is_active": True},
            {
                "_id": 0,
                "id": 1,
                "username": 1, 
                "role": 1,
                "first_name": 1,
                "last_name": 1,
                "is_first_login": 1
                # Exclude pin and pin_hash for security
            }
        ))
        
        # Sort admin users first
        users.sort(key=lambda x: (x.get("role") != "admin", x.get("username", "")))
        
        return users
    except Exception as e:
        print(f"Error fetching login users: {e}")
        # Return fallback admin user
        return [
            {
                "id": "admin-fallback",
                "username": "Admin",
                "role": "admin",
                "first_name": "System",
                "last_name": "Administrator",
                "is_first_login": True
            }
        ]

@app.get("/api/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    """Get all users (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = list(db.users.find({}, {"_id": 0, "pin_hash": 0}))
    return users

@app.post("/api/users")
async def create_user(user_data: dict, current_user: dict = Depends(get_current_user)):
    """Create new user (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if username already exists
    existing_user = db.users.find_one({"username": user_data["username"]})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Determine default PIN based on role
    role = user_data.get("role", "staff")
    if role == "admin":
        default_pin = "0000"
        requires_pin_change = False
    else:
        default_pin = "888888"  # Staff default PIN
        requires_pin_change = True  # Staff must change PIN on first login
    
    # Create new user
    new_user = User(
        id=str(uuid.uuid4()),
        username=user_data["username"],
        pin_hash=hash_pin(default_pin),
        role=role,
        email=user_data.get("email"),
        first_name=user_data.get("first_name"),
        last_name=user_data.get("last_name"),
        staff_id=user_data.get("staff_id"),
        created_at=datetime.utcnow(),
        is_first_login=requires_pin_change
    )
    
    db.users.insert_one(new_user.dict())
    
    # Remove sensitive data from response
    user_response = {k: v for k, v in new_user.dict().items() if k != "pin_hash"}
    user_response["default_pin"] = default_pin  # Include default PIN in response for admin
    return user_response

# Address autocomplete endpoint using free Nominatim API
@app.get("/api/address/search")
async def search_addresses(q: str, limit: int = 5):
    """Search addresses using OpenStreetMap Nominatim API (free)"""
    import httpx
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params={
                    "q": q,
                    "format": "json",
                    "addressdetails": 1,
                    "limit": limit,
                    "countrycodes": "au,us,gb,ca,nz"  # Limit to common English-speaking countries
                },
                headers={
                    "User-Agent": "RosterSync-AddressAutocomplete/1.0"
                }
            )
            
            if response.status_code == 200:
                results = response.json()
                formatted_results = []
                
                for result in results:
                    address = result.get("address", {})
                    formatted_results.append({
                        "display_name": result.get("display_name", ""),
                        "street_number": address.get("house_number", ""),
                        "route": address.get("road", ""),
                        "locality": address.get("city") or address.get("town") or address.get("village", ""),
                        "administrative_area_level_1": address.get("state", ""),
                        "country": address.get("country", ""),
                        "postal_code": address.get("postcode", ""),
                        "latitude": float(result.get("lat", 0)),
                        "longitude": float(result.get("lon", 0))
                    })
                
                return formatted_results
            else:
                return []
                
    except Exception as e:
        print(f"Address search error: {e}")
        return []

# Shift Request and Availability API Endpoints

@app.get("/api/unassigned-shifts")
async def get_unassigned_shifts(current_user: dict = Depends(get_current_user)):
    """Get all unassigned shifts (shifts without staff assigned)"""
    unassigned_shifts = list(db.roster.find({
        "$or": [
            {"staff_id": None},
            {"staff_id": ""},
            {"staff_name": None},
            {"staff_name": ""}
        ]
    }, {"_id": 0}))
    
    return sorted(unassigned_shifts, key=lambda x: (x['date'], x['start_time']))

@app.get("/api/shift-requests")
async def get_shift_requests(current_user: dict = Depends(get_current_user)):
    """Get shift requests - staff see their own, admin sees all"""
    query = {}
    if current_user["role"] == "staff":
        query["staff_id"] = current_user.get("staff_id", current_user["id"])
    
    shift_requests = list(db.shift_requests.find(query, {"_id": 0}))
    return sorted(shift_requests, key=lambda x: x.get('request_date', datetime.min), reverse=True)

@app.post("/api/shift-requests")
async def create_shift_request(request: ShiftRequest, current_user: dict = Depends(get_current_user)):
    """Create a new shift request"""
    if current_user["role"] != "staff":
        raise HTTPException(status_code=403, detail="Only staff can create shift requests")
    
    # Check if shift exists and is unassigned
    roster_entry = db.roster.find_one({"id": request.roster_entry_id})
    if not roster_entry:
        raise HTTPException(status_code=404, detail="Shift not found")
    
    if roster_entry.get("staff_id") or roster_entry.get("staff_name"):
        raise HTTPException(status_code=400, detail="Shift is already assigned")
    
    # Check for existing request from same staff for same shift
    existing_request = db.shift_requests.find_one({
        "roster_entry_id": request.roster_entry_id,
        "staff_id": current_user.get("staff_id", current_user["id"]),
        "status": {"$in": ["pending", "approved"]}
    })
    if existing_request:
        raise HTTPException(status_code=400, detail="You have already requested this shift")
    
    # Create the request
    request.id = str(uuid.uuid4())
    request.staff_id = current_user.get("staff_id", current_user["id"])
    request.staff_name = current_user.get("first_name") or current_user.get("username")
    request.request_date = datetime.utcnow()
    request.created_at = datetime.utcnow()
    
    db.shift_requests.insert_one(request.dict())
    return request

@app.put("/api/shift-requests/{request_id}/approve")
async def approve_shift_request(request_id: str, admin_notes: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Approve a shift request (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Find the request
    shift_request = db.shift_requests.find_one({"id": request_id})
    if not shift_request:
        raise HTTPException(status_code=404, detail="Shift request not found")
    
    if shift_request["status"] != "pending":
        raise HTTPException(status_code=400, detail="Request is not pending")
    
    # Check if shift is still unassigned
    roster_entry = db.roster.find_one({"id": shift_request["roster_entry_id"]})
    if not roster_entry:
        raise HTTPException(status_code=404, detail="Shift no longer exists")
    
    if roster_entry.get("staff_id") or roster_entry.get("staff_name"):
        raise HTTPException(status_code=400, detail="Shift has already been assigned")
    
    # Check for availability conflicts
    availability_conflicts = check_availability_conflicts(
        shift_request["staff_id"], 
        roster_entry["date"], 
        roster_entry["start_time"], 
        roster_entry["end_time"]
    )
    
    # Assign the shift
    db.roster.update_one(
        {"id": shift_request["roster_entry_id"]},
        {"$set": {
            "staff_id": shift_request["staff_id"],
            "staff_name": shift_request["staff_name"]
        }}
    )
    
    # Update request status
    db.shift_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "admin_notes": admin_notes,
            "approved_by": current_user["id"],
            "approved_date": datetime.utcnow()
        }}
    )
    
    # Create notification for staff
    notification = Notification(
        id=str(uuid.uuid4()),
        user_id=shift_request["staff_id"],
        notification_type=NotificationType.SHIFT_REQUEST_APPROVED,
        title="Shift Request Approved",
        message=f"Your request for shift on {roster_entry['date']} from {roster_entry['start_time']}-{roster_entry['end_time']} has been approved!",
        related_id=request_id,
        created_at=datetime.utcnow()
    )
    db.notifications.insert_one(notification.dict())
    
    return {"message": "Shift request approved successfully", "conflicts": availability_conflicts}

@app.put("/api/shift-requests/{request_id}/reject")
async def reject_shift_request(request_id: str, admin_notes: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Reject a shift request (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Find and update the request
    result = db.shift_requests.update_one(
        {"id": request_id, "status": "pending"},
        {"$set": {
            "status": "rejected",
            "admin_notes": admin_notes,
            "approved_by": current_user["id"],
            "approved_date": datetime.utcnow()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Shift request not found or not pending")
    
    # Get the request details for notification
    shift_request = db.shift_requests.find_one({"id": request_id})
    roster_entry = db.roster.find_one({"id": shift_request["roster_entry_id"]})
    
    # Create notification for staff
    notification = Notification(
        id=str(uuid.uuid4()),
        user_id=shift_request["staff_id"],
        notification_type=NotificationType.SHIFT_REQUEST_REJECTED,
        title="Shift Request Rejected",
        message=f"Your request for shift on {roster_entry['date'] if roster_entry else 'N/A'} has been rejected. {admin_notes or ''}",
        related_id=request_id,
        created_at=datetime.utcnow()
    )
    db.notifications.insert_one(notification.dict())
    
    return {"message": "Shift request rejected"}

# Additional CRUD endpoints for shift requests (Admin only)
@app.put("/api/shift-requests/{request_id}")
async def update_shift_request(request_id: str, request_update: ShiftRequest, current_user: dict = Depends(get_current_user)):
    """Update a shift request (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if request exists
    existing_request = db.shift_requests.find_one({"id": request_id})
    if not existing_request:
        raise HTTPException(status_code=404, detail="Shift request not found")
    
    # Update the request
    update_data = request_update.dict(exclude_unset=True)
    update_data["updated_at"] = datetime.utcnow()
    
    result = db.shift_requests.update_one(
        {"id": request_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Shift request not found")
    
    return {"message": "Shift request updated successfully"}

@app.delete("/api/shift-requests/{request_id}")
async def delete_shift_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a shift request (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if request exists
    existing_request = db.shift_requests.find_one({"id": request_id})
    if not existing_request:
        raise HTTPException(status_code=404, detail="Shift request not found")
    
    # Delete the request
    result = db.shift_requests.delete_one({"id": request_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Shift request not found")
    
    return {"message": "Shift request deleted successfully"}

@app.delete("/api/shift-requests")
async def clear_all_shift_requests(current_user: dict = Depends(get_current_user)):
    """Clear all shift requests (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Count existing requests
    existing_count = db.shift_requests.count_documents({})
    
    # Delete all requests
    result = db.shift_requests.delete_many({})
    
    return {
        "message": f"Cleared {result.deleted_count} shift requests successfully",
        "deleted_count": result.deleted_count
    }

@app.get("/api/staff-availability")
async def get_staff_availability(current_user: dict = Depends(get_current_user)):
    """Get staff availability - staff see their own, admin sees all"""
    query = {"is_active": True}
    if current_user["role"] == "staff":
        query["staff_id"] = current_user.get("staff_id", current_user["id"])
    
    availability_records = list(db.staff_availability.find(query, {"_id": 0}))
    return sorted(availability_records, key=lambda x: x.get('created_at', datetime.min), reverse=True)

@app.post("/api/staff-availability")
async def create_staff_availability(availability: StaffAvailability, current_user: dict = Depends(get_current_user)):
    """Create staff availability record"""
    # Staff can only create their own records
    if current_user["role"] == "staff":
        availability.staff_id = current_user.get("staff_id", current_user["id"])
        availability.staff_name = current_user.get("first_name") or current_user.get("username")
    else:
        # Admin users must provide valid staff_id and staff_name
        if not availability.staff_id or not availability.staff_name:
            raise HTTPException(
                status_code=422, 
                detail="Admin must provide staff_id and staff_name when creating availability"
            )
        
        # Validate that the staff member exists
        staff_member = db.staff.find_one({"id": availability.staff_id, "active": True})
        if not staff_member:
            raise HTTPException(
                status_code=404, 
                detail="Staff member not found or inactive"
            )
    
    # Additional validation for all users
    if not availability.staff_id or not availability.staff_name:
        raise HTTPException(
            status_code=422, 
            detail="staff_id and staff_name are required fields"
        )
    
    availability.id = str(uuid.uuid4())
    availability.created_at = datetime.utcnow()
    
    db.staff_availability.insert_one(availability.dict())
    return availability

@app.put("/api/staff-availability/{availability_id}")
async def update_staff_availability(availability_id: str, availability: StaffAvailability, current_user: dict = Depends(get_current_user)):
    """Update staff availability record"""
    # Check if record exists and user has permission
    existing_record = db.staff_availability.find_one({"id": availability_id})
    if not existing_record:
        raise HTTPException(status_code=404, detail="Availability record not found")
    
    # Staff can only update their own records
    if current_user["role"] == "staff" and existing_record["staff_id"] != current_user.get("staff_id", current_user["id"]):
        raise HTTPException(status_code=403, detail="You can only update your own availability")
    
    result = db.staff_availability.update_one(
        {"id": availability_id},
        {"$set": availability.dict()}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Availability record not found")
    
    return availability

@app.delete("/api/staff-availability/{availability_id}")
async def delete_staff_availability(availability_id: str, current_user: dict = Depends(get_current_user)):
    """Delete staff availability record"""
    existing_record = db.staff_availability.find_one({"id": availability_id})
    if not existing_record:
        raise HTTPException(status_code=404, detail="Availability record not found")
    
    # Staff can only delete their own records
    if current_user["role"] == "staff" and existing_record["staff_id"] != current_user.get("staff_id", current_user["id"]):
        raise HTTPException(status_code=403, detail="You can only delete your own availability")
    
    result = db.staff_availability.update_one(
        {"id": availability_id},
        {"$set": {"is_active": False}}
    )
    
    return {"message": "Availability record deleted"}

@app.delete("/api/staff-availability")
async def clear_all_staff_availability(current_user: dict = Depends(get_current_user)):
    """Clear all staff availability records (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Count existing active records
    existing_count = db.staff_availability.count_documents({"is_active": True})
    
    # Soft delete all availability records
    result = db.staff_availability.update_many(
        {"is_active": True},
        {"$set": {"is_active": False, "deleted_at": datetime.utcnow()}}
    )
    
    return {
        "message": f"Cleared {result.modified_count} staff availability records successfully",
        "cleared_count": result.modified_count
    }

@app.post("/api/admin/sync_staff_users")
async def sync_staff_users(current_user: dict = Depends(get_current_user)):
    """Create missing user accounts for all active staff members (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get all active staff members
    staff_members = list(db.staff.find({"active": True}))
    
    created_users = []
    existing_users = []
    errors = []
    
    for staff in staff_members:
        # Skip staff with empty names
        if not staff.get("name") or not staff.get("name").strip():
            errors.append(f"Skipped staff with empty name (ID: {staff.get('id')})")
            continue
        
        # Generate username from staff name
        staff_name = staff["name"].strip()
        username = staff_name.lower().replace(' ', '')
        
        # Check if user already exists
        existing_user = db.users.find_one({"username": username})
        if existing_user:
            existing_users.append(f"{staff_name} -> {username}")
            continue
        
        try:
            # Create new user account
            user_id = str(uuid.uuid4())
            default_pin = "888888"
            pin_hash = hash_pin(default_pin)
            
            new_user = User(
                id=user_id,
                username=username,
                pin_hash=pin_hash,
                role=UserRole.STAFF,
                first_name=staff_name.split()[0] if staff_name.split() else staff_name,
                last_name=' '.join(staff_name.split()[1:]) if len(staff_name.split()) > 1 else '',
                email=f"{username}@company.com",
                staff_id=staff["id"],
                is_active=True,
                is_first_login=True,  # Force PIN change on first login
                created_at=datetime.utcnow()
            )
            
            db.users.insert_one(new_user.dict())
            created_users.append(f"{staff_name} -> {username} (PIN: {default_pin})")
            
        except Exception as e:
            errors.append(f"Failed to create user for {staff_name}: {str(e)}")
    
    # Clean up staff with empty names
    empty_name_staff = list(db.staff.find({"$or": [{"name": ""}, {"name": None}]}))
    cleaned_up = []
    for empty_staff in empty_name_staff:
        db.staff.update_one({"id": empty_staff["id"]}, {"$set": {"active": False}})
        cleaned_up.append(empty_staff["id"])
    
    result = {
        "message": f"Staff user synchronization completed",
        "created_users": created_users,
        "existing_users": existing_users,
        "errors": errors,
        "cleaned_up_empty_names": cleaned_up,
        "summary": {
            "created": len(created_users),
            "existing": len(existing_users),
            "errors": len(errors),
            "cleaned_up": len(cleaned_up)
        }
    }
    
    return result

@app.get("/api/notifications")
async def get_notifications(current_user: dict = Depends(get_current_user)):
    """Get notifications for current user"""
    notifications = list(db.notifications.find(
        {"user_id": current_user["id"]},
        {"_id": 0}
    ))
    return sorted(notifications, key=lambda x: x.get('created_at', datetime.min), reverse=True)

@app.put("/api/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Mark notification as read"""
    result = db.notifications.update_one(
        {"id": notification_id, "user_id": current_user["id"]},
        {"$set": {"is_read": True}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification marked as read"}

def check_availability_conflicts(staff_id: str, date: str, start_time: str, end_time: str) -> List[Dict]:
    """Check for availability conflicts when assigning a shift"""
    conflicts = []
    
    # Convert times to minutes for comparison
    start_minutes = int(start_time.split(':')[0]) * 60 + int(start_time.split(':')[1])
    end_minutes = int(end_time.split(':')[0]) * 60 + int(end_time.split(':')[1])
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60  # Handle overnight shifts
    
    # Parse date
    shift_date = datetime.strptime(date, "%Y-%m-%d")
    day_of_week = shift_date.weekday()
    
    # Check for unavailability or time off requests
    unavailable_records = list(db.staff_availability.find({
        "staff_id": staff_id,
        "is_active": True,
        "availability_type": {"$in": ["unavailable", "time_off_request"]},
        "$or": [
            # Specific date match
            {"date_from": {"$lte": date}, "date_to": {"$gte": date}},
            # Single date match
            {"date_from": date, "date_to": None},
            # Recurring weekly unavailability
            {"is_recurring": True, "day_of_week": day_of_week}
        ]
    }))
    
    for record in unavailable_records:
        conflict_times = []
        
        # Check time overlap if times are specified
        if record.get("start_time") and record.get("end_time"):
            rec_start = int(record["start_time"].split(':')[0]) * 60 + int(record["start_time"].split(':')[1])
            rec_end = int(record["end_time"].split(':')[0]) * 60 + int(record["end_time"].split(':')[1])
            if rec_end <= rec_start:
                rec_end += 24 * 60
            
            # Check for time overlap
            if not (end_minutes <= rec_start or start_minutes >= rec_end):
                conflict_times = [record["start_time"], record["end_time"]]
        else:
            # All day unavailability
            conflict_times = ["All Day"]
        
        if conflict_times or not record.get("start_time"):
            conflicts.append({
                "type": record["availability_type"],
                "date_range": f"{record.get('date_from', date)} to {record.get('date_to', date)}",
                "times": conflict_times,
                "notes": record.get("notes", ""),
                "is_recurring": record.get("is_recurring", False)
            })
    
    return conflicts

@app.post("/api/check-assignment-conflicts")
async def check_assignment_conflicts(data: dict, current_user: dict = Depends(get_current_user)):
    """Check for conflicts when admin tries to assign staff to a shift"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    staff_id = data.get("staff_id")
    date = data.get("date")
    start_time = data.get("start_time")
    end_time = data.get("end_time")
    
    if not all([staff_id, date, start_time, end_time]):
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    conflicts = check_availability_conflicts(staff_id, date, start_time, end_time)
    
    return {
        "has_conflicts": len(conflicts) > 0,
        "conflicts": conflicts,
        "can_override": True,  # Admin can always override
        "message": f"Found {len(conflicts)} potential conflicts" if conflicts else "No conflicts found"
    }

# ===========================================
# CLIENT PROFILE MANAGEMENT ENDPOINTS
# ===========================================

@app.get("/api/clients")
async def get_clients(current_user: dict = Depends(get_current_user)):
    """Get all client profiles with role-based filtering"""
    
    # For staff, return limited information (basic details only, no NDIS plan details)
    if current_user["role"] == "staff":
        clients = list(db.clients.find({}, {
            "_id": 0,
            "id": 1,
            "full_name": 1,
            "date_of_birth": 1,
            "age": 1,
            "sex": 1,
            "disability_condition": 1,
            "mobile": 1,
            "address": 1,
            "emergency_contacts": 1,
            "biography.strengths": 1,
            "biography.daily_life": 1,
            "biography.additional_info": 1,
            "created_at": 1
        }))
    else:
        # Admin and Supervisor get full access
        clients = list(db.clients.find({}, {"_id": 0}))
    
    return sorted(clients, key=lambda x: x.get('full_name', ''))

@app.get("/api/clients/{client_id}")
async def get_client_profile(client_id: str, current_user: dict = Depends(get_current_user)):
    """Get specific client profile with role-based filtering"""
    
    if current_user["role"] == "staff":
        # Staff get limited view (no NDIS plan financial details, limited biography)
        client = db.clients.find_one({"id": client_id}, {
            "_id": 0,
            "id": 1,
            "full_name": 1,
            "date_of_birth": 1,
            "age": 1,
            "sex": 1,
            "disability_condition": 1,
            "mobile": 1,
            "address": 1,
            "emergency_contacts": 1,
            "biography.strengths": 1,
            "biography.daily_life": 1,
            "biography.additional_info": 1,
            "created_at": 1,
            "current_ndis_plan.plan_type": 1,
            "current_ndis_plan.ndis_number": 1,
            "current_ndis_plan.plan_start_date": 1,
            "current_ndis_plan.plan_end_date": 1
        })
    else:
        # Full access for Admin and Supervisor
        client = db.clients.find_one({"id": client_id}, {"_id": 0})
    
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return client

@app.post("/api/clients")
async def create_client_profile(client: ClientProfile, current_user: dict = Depends(get_current_user)):
    """Create new client profile (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Generate ID and set creation fields
    client.id = str(uuid.uuid4())
    client.created_by = current_user["username"]
    client.created_at = datetime.utcnow()
    client.updated_at = datetime.utcnow()
    
    # Calculate age if date of birth provided
    if client.date_of_birth:
        try:
            birth_date = datetime.strptime(client.date_of_birth, "%d/%m/%Y")
            today = datetime.now()
            client.age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        except ValueError:
            # If date format is different, skip age calculation
            pass
    
    # Insert client profile
    db.clients.insert_one(client.dict())
    return client

@app.put("/api/clients/{client_id}")
async def update_client_profile(client_id: str, client_update: ClientProfile, current_user: dict = Depends(get_current_user)):
    """Update client profile (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if client exists
    existing_client = db.clients.find_one({"id": client_id})
    if not existing_client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Update fields
    update_data = client_update.dict(exclude_unset=True)
    update_data["updated_at"] = datetime.utcnow()
    
    # Recalculate age if date of birth is updated
    if "date_of_birth" in update_data:
        try:
            birth_date = datetime.strptime(update_data["date_of_birth"], "%d/%m/%Y")
            today = datetime.now()
            update_data["age"] = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        except ValueError:
            pass
    
    result = db.clients.update_one(
        {"id": client_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return {"message": "Client profile updated successfully"}

@app.delete("/api/clients/{client_id}")
async def delete_client_profile(client_id: str, current_user: dict = Depends(get_current_user)):
    """Delete client profile (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if client exists
    existing_client = db.clients.find_one({"id": client_id})
    if not existing_client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Soft delete (mark as inactive instead of removing)
    result = db.clients.update_one(
        {"id": client_id},
        {"$set": {"is_active": False, "deleted_at": datetime.utcnow()}}
    )
    
    return {"message": "Client profile deleted successfully"}

@app.post("/api/clients/{client_id}/ndis-plan")
async def update_ndis_plan(client_id: str, ndis_plan: NDISPlan, current_user: dict = Depends(get_current_user)):
    """Update NDIS plan for client (Admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if client exists
    existing_client = db.clients.find_one({"id": client_id})
    if not existing_client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Calculate remaining amounts for funding categories
    for category in ndis_plan.funding_categories:
        category.remaining_amount = category.total_amount - category.spent_amount
    
    # Update NDIS plan
    result = db.clients.update_one(
        {"id": client_id},
        {
            "$set": {
                "current_ndis_plan": ndis_plan.dict(),
                "updated_at": datetime.utcnow()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return {"message": "NDIS plan updated successfully"}

@app.get("/api/clients/{client_id}/budget-summary")
async def get_client_budget_summary(client_id: str, current_user: dict = Depends(get_current_user)):
    """Get budget summary for client's NDIS plan (Admin/Supervisor only)"""
    if current_user["role"] == "staff":
        raise HTTPException(status_code=403, detail="Access denied - insufficient permissions")
    
    client = db.clients.find_one({"id": client_id})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    ndis_plan = client.get("current_ndis_plan")
    if not ndis_plan:
        return {"message": "No NDIS plan found for this client"}
    
    # Calculate budget summary
    budget_summary = {
        "plan_overview": {
            "ndis_number": ndis_plan.get("ndis_number"),
            "plan_start": ndis_plan.get("plan_start_date"),
            "plan_end": ndis_plan.get("plan_end_date"),
            "plan_management": ndis_plan.get("plan_management")
        },
        "funding_summary": {
            "total_funding": 0,
            "total_spent": 0,
            "total_remaining": 0,
            "categories": []
        }
    }
    
    for category in ndis_plan.get("funding_categories", []):
        category_data = {
            "category_name": category.get("category_name"),
            "total_amount": category.get("total_amount", 0),
            "spent_amount": category.get("spent_amount", 0),
            "remaining_amount": category.get("remaining_amount", category.get("total_amount", 0) - category.get("spent_amount", 0)),
            "utilization_percentage": round((category.get("spent_amount", 0) / category.get("total_amount", 1)) * 100, 2) if category.get("total_amount", 0) > 0 else 0,
            "funding_period": category.get("funding_period")
        }
        
        budget_summary["funding_summary"]["categories"].append(category_data)
        budget_summary["funding_summary"]["total_funding"] += category.get("total_amount", 0)
        budget_summary["funding_summary"]["total_spent"] += category.get("spent_amount", 0)
        budget_summary["funding_summary"]["total_remaining"] += category_data["remaining_amount"]
    
    # Overall utilization percentage
    total_funding = budget_summary["funding_summary"]["total_funding"]
    budget_summary["funding_summary"]["overall_utilization"] = round(
        (budget_summary["funding_summary"]["total_spent"] / total_funding) * 100, 2
    ) if total_funding > 0 else 0
    
    return budget_summary

@app.put("/api/clients/{client_id}/biography")
async def update_client_biography(client_id: str, biography_update: ClientBiography, current_user: dict = Depends(get_current_user)):
    """Update client biography information with role-based access"""
    
    # Check if client exists
    existing_client = db.clients.find_one({"id": client_id})
    if not existing_client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Role-based access control
    if current_user["role"] == "staff":
        # Staff can only update certain fields (not supports or funding-related info)
        allowed_fields = ["strengths", "daily_life", "additional_info"]
        biography_data = biography_update.dict()
        filtered_data = {k: v for k, v in biography_data.items() if k in allowed_fields}
        
        if not filtered_data:
            raise HTTPException(status_code=403, detail="Staff users have limited biography editing permissions")
        
        update_data = {"biography": filtered_data, "updated_at": datetime.utcnow()}
        
        # Merge with existing biography data to preserve admin-only fields
        existing_bio = existing_client.get("biography", {})
        merged_bio = {**existing_bio, **filtered_data}
        update_data["biography"] = merged_bio
    else:
        # Admin and Supervisor get full access
        update_data = {
            "biography": biography_update.dict(),
            "updated_at": datetime.utcnow()
        }
    
    result = db.clients.update_one(
        {"id": client_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    return {"message": "Client biography updated successfully"}

# Initialize database with sample client if not exists (for demonstration)
def initialize_sample_client():
    """Initialize sample client data if not exists"""
    sample_client = db.clients.find_one({"full_name": "Jeremy James Tomlinson"})
    if not sample_client:
        print("Creating sample client profile...")
        
        # Create Jeremy's emergency contacts
        emergency_contacts = [
            EmergencyContact(
                name="Jo-ann Tomlinson",
                relationship="Mother",
                mobile="0412110888",
                address="21 Dialba CRS, Tingalpa 4173, Brisbane, QLD"
            ),
            EmergencyContact(
                name="Brett Tomlinson",
                relationship="Father",
                mobile="0448066351",
                address="21 Dialba CRS, Tingalpa 4173, Brisbane, QLD"
            )
        ]
        
        # Create plan manager details
        plan_manager = PlanManagerDetails(
            provider_name="Charter Care Plan Management",
            contact_person="Kirsty Condon",
            phone="0474464163",
            email="kirsty.condon@ccpms.au"
        )
        
        # Create funding categories
        funding_categories = [
            NDISFundingCategory(
                category_name="Consumables",
                total_amount=3167.81,
                funding_period=FundingPeriod.QUARTERLY,
                description="Continence aid total $2,167.81 for a year, $1000 for low cost AT",
                spent_amount=0.0
            ),
            NDISFundingCategory(
                category_name="Home and Living",
                total_amount=885331.00,
                funding_period=FundingPeriod.MONTHLY,
                description="$865,798.34 for regular SIL supports, $19,532.66 for irregular SIL supports",
                spent_amount=0.0
            ),
            NDISFundingCategory(
                category_name="Social Community and Civic Participation",
                total_amount=67101.10,
                funding_period=FundingPeriod.QUARTERLY,
                description="2 hours per day high intensity 1:1 support including weekend and public holiday",
                spent_amount=0.0
            ),
            NDISFundingCategory(
                category_name="Transport",
                total_amount=1784.00,
                funding_period=FundingPeriod.QUARTERLY,
                description="Support to access community activities",
                spent_amount=0.0
            ),
            NDISFundingCategory(
                category_name="Support Coordination",
                total_amount=21251.60,
                funding_period=FundingPeriod.QUARTERLY,
                description="Level 2: 60 hours ($6,008.40), Level 3: 80 hours ($15,243.20)",
                spent_amount=0.0
            ),
            NDISFundingCategory(
                category_name="Improved Daily Living",
                total_amount=48999.83,
                funding_period=FundingPeriod.QUARTERLY,
                description="OT, Physiotherapy, Dietitian, Speech pathology, Psychology, Therapy assistance, Registered nurse",
                spent_amount=0.0
            ),
            NDISFundingCategory(
                category_name="Assistive Technology",
                total_amount=5050.00,
                funding_period=FundingPeriod.QUARTERLY,
                description="Maintenance, repair and rental of AT equipment",
                spent_amount=0.0
            ),
            NDISFundingCategory(
                category_name="Specialist Disability Accommodation (SDA)",
                total_amount=78660.00,
                funding_period=FundingPeriod.MONTHLY,
                description="High physical support, Apartment, two bedrooms, Brisbane South",
                spent_amount=0.0
            )
        ]
        
        # Create NDIS plan
        ndis_plan = NDISPlan(
            plan_type="PACE",
            ndis_number="431525388",
            plan_start_date="31/07/2025",
            plan_end_date="30/07/2026",
            plan_management=PlanManagement.PLAN_MANAGED,
            plan_manager=plan_manager,
            funding_categories=funding_categories
        )
        
        # Create client profile
        sample_client = ClientProfile(
            id=str(uuid.uuid4()),
            full_name="Jeremy James Tomlinson",
            date_of_birth="17/09/1988",
            age=36,  # Will be calculated automatically
            sex="Male",
            disability_condition="Refractory Polymyositis, FSHD Type 2",
            mobile="0490821919",
            address="20/384 Stanley Rd, Carina 4152, Brisbane, QLD",
            emergency_contacts=emergency_contacts,
            current_ndis_plan=ndis_plan,
            created_by="System",
            created_at=datetime.utcnow()
        )
        
        db.clients.insert_one(sample_client.dict())
        print("✅ Sample client profile created: Jeremy James Tomlinson")
    else:
        print("✅ Sample client already exists")

# =====================================
# OCR DOCUMENT PROCESSING ENDPOINTS
# =====================================

@app.post("/api/ocr/process")
async def process_document(
    file: UploadFile = File(...),
    client_id: str = Form(None),
    extract_client_data: bool = Form(True),
    current_user: dict = Depends(get_current_user)
):
    """Process uploaded document and extract text using OCR"""
    
    # Check user permissions (Admin and Supervisor only)
    if current_user.get("role") not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="Not authorized to process documents")
    
    # Generate unique task ID
    task_id = str(uuid.uuid4())
    
    logging.info(f"🔄 Starting OCR processing for {file.filename} (task: {task_id}, user: {current_user.get('username')})")
    
    try:
        # Validate and save file
        file_path = await validate_and_save_file(file)
        logging.info(f"✅ File saved successfully: {file_path}")
        
        # Initialize processing status
        ocr_results[task_id] = {
            'id': task_id,
            'client_id': client_id,
            'filename': file.filename,
            'file_type': file.content_type,
            'status': 'processing',
            'progress': 0,
            'created_at': datetime.now().isoformat(),
            'created_by': current_user["username"]
        }
        
        logging.info(f"🚀 Processing document: {file.filename} (type: {file.content_type})")
        
        # Process document based on type
        if file.content_type == 'application/pdf' or file.filename.lower().endswith('.pdf'):
            logging.info(f"📄 Processing as PDF: {file.filename}")
            result = await process_pdf_async(file_path, task_id)
        else:
            logging.info(f"🖼️ Processing as image: {file.filename}")
            result = await process_image_async(file_path, task_id)
        
        # Extract text for parsing
        extracted_text = result.get('combined_text', result.get('text', ''))
        logging.info(f"📝 Extracted {len(extracted_text)} characters of text from {file.filename}")
        
        # Parse NDIS plan data if requested
        extracted_data = None
        if extract_client_data and extracted_text:
            logging.info(f"🔍 Parsing NDIS plan data from {file.filename}")
            extracted_data = ocr_processor.parse_ndis_plan_text(extracted_text)
        
        # Update results
        ocr_results[task_id].update({
            'status': 'completed',
            'result': result,
            'extracted_text': extracted_text,
            'extracted_data': extracted_data.dict() if extracted_data else None,
            'completed_at': datetime.now().isoformat(),
            'progress': 100
        })
        
        logging.info(f"✅ OCR processing completed successfully for {file.filename}")
        
        # Clean up file
        file_path.unlink(missing_ok=True)
        
        return JSONResponse({
            'task_id': task_id,
            'status': 'completed',
            'message': 'Document processing completed successfully',
            'extracted_data': extracted_data.dict() if extracted_data else None
        })
        
    except HTTPException as e:
        logging.error(f"❌ HTTP Exception during OCR processing: {e.detail}")
        raise
    except Exception as e:
        error_msg = str(e)
        logging.error(f"❌ Critical error during OCR processing for {file.filename}: {error_msg}")
        ocr_results[task_id] = {
            'id': task_id,
            'filename': file.filename,
            'status': 'failed',
            'error': error_msg,
            'created_at': datetime.now().isoformat(),
            'created_by': current_user["username"]
        }
        raise HTTPException(status_code=500, detail=f"Processing failed: {error_msg}")

@app.get("/api/ocr/status/{task_id}")
async def get_processing_status(task_id: str, current_user: dict = Depends(get_current_user)):
    """Get the processing status of a document"""
    
    # Check user permissions
    if current_user.get("role") not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="Not authorized to view OCR results")
    
    if task_id not in ocr_results:
        raise HTTPException(status_code=404, detail="Task not found")
    
    return ocr_results[task_id]

@app.get("/api/ocr/result/{task_id}")
async def get_processing_result(task_id: str, current_user: dict = Depends(get_current_user)):
    """Get the OCR result for a completed task"""
    
    # Check user permissions
    if current_user.get("role") not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="Not authorized to view OCR results")
    
    if task_id not in ocr_results:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task_data = ocr_results[task_id]
    
    if task_data['status'] == 'processing':
        raise HTTPException(status_code=202, detail="Processing still in progress")
    
    if task_data['status'] == 'failed':
        raise HTTPException(status_code=500, detail=task_data.get('error', 'Processing failed'))
    
    return {
        'task_id': task_id,
        'status': task_data['status'],
        'result': task_data.get('result', {}),
        'extracted_text': task_data.get('extracted_text', ''),
        'extracted_data': task_data.get('extracted_data', {}),
        'processing_info': {
            'filename': task_data['filename'],
            'file_type': task_data['file_type'],
            'created_at': task_data['created_at'],
            'completed_at': task_data.get('completed_at'),
            'created_by': task_data.get('created_by')
        }
    }

@app.post("/api/ocr/apply-to-client/{task_id}")
async def apply_ocr_to_client(
    task_id: str, 
    client_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Apply extracted OCR data to a client profile"""
    
    # Check user permissions (Admin only for now)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized to modify client profiles")
    
    if task_id not in ocr_results:
        raise HTTPException(status_code=404, detail="OCR task not found")
    
    task_data = ocr_results[task_id]
    
    if task_data['status'] != 'completed':
        raise HTTPException(status_code=400, detail="OCR processing not completed")
    
    extracted_data = task_data.get('extracted_data')
    if not extracted_data:
        raise HTTPException(status_code=400, detail="No extracted data available")
    
    try:
        if client_id:
            # Update existing client
            existing_client = db.clients.find_one({"id": client_id})
            if not existing_client:
                raise HTTPException(status_code=404, detail="Client not found")
            
            # Prepare update data (only update empty fields)
            update_data = {}
            if not existing_client.get('full_name') and extracted_data.get('full_name'):
                update_data['full_name'] = extracted_data['full_name']
            if not existing_client.get('date_of_birth') and extracted_data.get('date_of_birth'):
                update_data['date_of_birth'] = extracted_data['date_of_birth']
            if not existing_client.get('disability_condition') and extracted_data.get('disability_condition'):
                update_data['disability_condition'] = extracted_data['disability_condition']
            if not existing_client.get('address') and extracted_data.get('address'):
                update_data['address'] = extracted_data['address']
            if not existing_client.get('mobile') and extracted_data.get('mobile'):
                update_data['mobile'] = extracted_data['mobile']
            
            # Update NDIS plan information
            if not existing_client.get('current_ndis_plan'):
                ndis_plan = {
                    "ndis_number": extracted_data.get('ndis_number'),
                    "plan_start_date": extracted_data.get('plan_start_date'),
                    "plan_end_date": extracted_data.get('plan_end_date'),
                    "plan_type": "PACE",  # Default
                    "plan_management": "self_managed",  # Default
                    "funding_categories": extracted_data.get('funding_categories', []),
                    "plan_document_path": task_data.get('filename'),
                    "created_at": datetime.now(),
                    "updated_at": datetime.now()
                }
                update_data['current_ndis_plan'] = ndis_plan
            
            if update_data:
                update_data['updated_at'] = datetime.now()
                db.clients.update_one({"id": client_id}, {"$set": update_data})
                
                return {
                    "message": "Client profile updated successfully with OCR data",
                    "client_id": client_id,
                    "updated_fields": list(update_data.keys())
                }
            else:
                return {
                    "message": "No updates needed - client profile already complete",
                    "client_id": client_id
                }
        else:
            # Create new client profile
            new_client_id = str(uuid.uuid4())
            
            client_data = {
                "id": new_client_id,
                "full_name": extracted_data.get('full_name', ''),
                "date_of_birth": extracted_data.get('date_of_birth', ''),
                "sex": "Not Specified",  # Default
                "disability_condition": extracted_data.get('disability_condition', ''),
                "mobile": extracted_data.get('mobile', ''),
                "address": extracted_data.get('address', ''),
                "emergency_contacts": [],
                "current_ndis_plan": {
                    "ndis_number": extracted_data.get('ndis_number'),
                    "plan_start_date": extracted_data.get('plan_start_date'),
                    "plan_end_date": extracted_data.get('plan_end_date'),
                    "plan_type": "PACE",  # Default
                    "plan_management": "self_managed",  # Default
                    "funding_categories": extracted_data.get('funding_categories', []),
                    "plan_document_path": task_data.get('filename'),
                    "created_at": datetime.now(),
                    "updated_at": datetime.now()
                },
                "created_at": datetime.now(),
                "updated_at": datetime.now(),
                "created_by": current_user["username"]
            }
            
            db.clients.insert_one(client_data)
            
            return {
                "message": "New client profile created successfully from OCR data",
                "client_id": new_client_id,
                "confidence_score": extracted_data.get('confidence_score', 0)
            }
    
    except Exception as e:
        logging.error(f"Error applying OCR data to client: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to apply OCR data: {str(e)}")

@app.get("/api/ocr/health")
async def ocr_health_check():
    """Health check for OCR service"""
    try:
        # Test Tesseract installation
        version = pytesseract.get_tesseract_version()
        return {
            'status': 'healthy',
            'tesseract_version': str(version),
            'upload_dir_exists': UPLOAD_DIR.exists(),
            'timestamp': datetime.now().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"OCR service unhealthy: {str(e)}")

@app.delete("/api/ocr/cleanup")
async def cleanup_ocr_results(current_user: dict = Depends(get_current_user)):
    """Clean up old OCR results (Admin only)"""
    
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized to cleanup OCR data")
    
    # Remove results older than 24 hours
    cutoff_time = datetime.now() - timedelta(hours=24)
    
    cleaned_count = 0
    task_ids_to_remove = []
    
    for task_id, task_data in ocr_results.items():
        try:
            created_at = datetime.fromisoformat(task_data['created_at'])
            if created_at < cutoff_time:
                task_ids_to_remove.append(task_id)
                cleaned_count += 1
        except Exception:
            # If we can't parse the date, remove it anyway
            task_ids_to_remove.append(task_id)
            cleaned_count += 1
    
    for task_id in task_ids_to_remove:
        del ocr_results[task_id]
    
    return {
        "message": f"Cleaned up {cleaned_count} old OCR results",
        "cleaned_count": cleaned_count
    }

# =====================================
# EXPORT FUNCTIONALITY ENDPOINTS
# =====================================

def get_roster_data_for_export(start_date: str, end_date: str, current_user: dict):
    """Get roster data for export with date range support"""
    
    # Query roster entries for the date range
    query = {
        "date": {"$gte": start_date, "$lt": end_date}
    }
    
    roster_entries = list(db.roster.find(query, {"_id": 0}).sort("date", 1))
    
    # Get staff information
    staff_dict = {}
    for staff in db.staff.find({}, {"_id": 0}):
        staff_dict[staff["id"]] = staff
    
    # Get settings for rates
    settings = db.settings.find_one() or {}
    rates = settings.get("rates", {})
    
    # Process entries and add calculated information
    export_data = []
    for entry in roster_entries:
        staff_info = staff_dict.get(entry.get("staff_id", ""), {})
        staff_name = staff_info.get("name", "Unassigned") if entry.get("staff_id") else "Unassigned"
        
        # Role-based data filtering
        if current_user["role"] == "staff":
            # Staff can only see their own shifts
            staff_id = current_user.get("staff_id") or current_user.get("id")
            if entry.get("staff_id") != staff_id:
                continue
        
        # Calculate pay information
        hours_worked = entry.get("hours_worked", 0)
        base_pay = entry.get("base_pay", 0)
        total_pay = entry.get("total_pay", 0)
        
        # Calculate hourly rate from base pay and hours worked
        if hours_worked > 0 and not entry.get("is_sleepover", False):
            hourly_rate = base_pay / hours_worked
        else:
            hourly_rate = 0
        
        export_entry = {
            "Date": entry.get("date", ""),
            "Day of Week": datetime.strptime(entry.get("date", ""), "%Y-%m-%d").strftime("%A") if entry.get("date") else "",
            "Staff Name": staff_name,
            "Start Time": entry.get("start_time", ""),
            "End Time": entry.get("end_time", ""),
            "Hours Worked": f"{hours_worked:.1f}h",
            "Shift Type": entry.get("shift_type", "").replace("_", " ").title(),
            "Is Sleepover": "Yes" if entry.get("is_sleepover", False) else "No",
            "Hourly Rate": f"${hourly_rate:.2f}" if not entry.get("is_sleepover") else "Sleepover",
            "Total Pay": f"${total_pay:.2f}",
            "Client": entry.get("client_name", "Unassigned"),
            "Location": entry.get("location", ""),
            "Notes": entry.get("notes", "")
        }
        
        # Add NDIS information for admin users only
        if current_user["role"] in ["admin", "supervisor"]:
            export_entry.update({
                "NDIS Hourly Charge": f"${entry.get('ndis_hourly_charge', 0):.2f}",
                "NDIS Total Charge": f"${entry.get('ndis_total_charge', 0):.2f}",
                "NDIS Line Item": entry.get('ndis_line_item_code', ''),
                "NDIS Description": entry.get('ndis_description', '')
            })
        
        export_data.append(export_entry)
    
    return export_data

@app.get("/api/export/csv/{month}")
async def export_csv(month: str, current_user: dict = Depends(get_current_user)):
    """Export roster data as CSV file"""
    
    try:
        # Parse month (YYYY-MM format) and convert to date range
        try:
            year, month_num = map(int, month.split("-"))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        
        start_date = f"{year}-{month_num:02d}-01"
        if month_num == 12:
            end_year = year + 1
            end_month = 1
        else:
            end_year = year
            end_month = month_num + 1
        end_date = f"{end_year}-{end_month:02d}-01"
        
        # Get roster data
        export_data = get_roster_data_for_export(start_date, end_date, current_user)
        
        if not export_data:
            raise HTTPException(status_code=404, detail="No roster data found for the specified month")
        
        # Create CSV in memory
        output = io.StringIO()
        if export_data:
            fieldnames = export_data[0].keys()
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(export_data)
        
        # Create response
        csv_content = output.getvalue()
        output.close()
        
        # Convert to bytes for streaming response
        csv_bytes = io.BytesIO(csv_content.encode('utf-8'))
        
        month_name = datetime.strptime(f"{month}-01", "%Y-%m-%d").strftime("%B_%Y")
        filename = f"roster_export_{month_name}.csv"
        
        return StreamingResponse(
            io.BytesIO(csv_content.encode('utf-8')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException as he:
        # Re-raise HTTP exceptions (like 404) without wrapping them
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating CSV: {str(e)}")

@app.get("/api/export/excel/{month}")
async def export_excel(month: str, current_user: dict = Depends(get_current_user)):
    """Export roster data as Excel file"""
    
    try:
        # Parse month (YYYY-MM format) and convert to date range
        try:
            year, month_num = map(int, month.split("-"))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        
        start_date = f"{year}-{month_num:02d}-01"
        if month_num == 12:
            end_year = year + 1
            end_month = 1
        else:
            end_year = year
            end_month = month_num + 1
        end_date = f"{end_year}-{end_month:02d}-01"
        
        # Get roster data
        export_data = get_roster_data_for_export(start_date, end_date, current_user)
        
        if not export_data:
            raise HTTPException(status_code=404, detail="No roster data found for the specified month")
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        # Create workbook and worksheet
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Convert data to DataFrame
            df = pd.DataFrame(export_data)
            
            # Write to Excel
            sheet_name = f"Roster {month}"
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get the workbook and worksheet for styling
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Style the header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column width
                column_letter = openpyxl.utils.get_column_letter(col_num)
                worksheet.column_dimensions[column_letter].width = max(len(str(column_title)) + 2, 12)
            
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Data rows
                        cell.alignment = Alignment(horizontal="center")
        
        output.seek(0)
        
        month_name = datetime.strptime(f"{month}-01", "%Y-%m-%d").strftime("%B_%Y")
        filename = f"roster_export_{month_name}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException as he:
        # Re-raise HTTP exceptions (like 404) without wrapping them
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")

@app.get("/api/export/pdf/{month}")
async def export_pdf(month: str, current_user: dict = Depends(get_current_user)):
    """Export roster data as PDF file"""
    
    try:
        # Parse month (YYYY-MM format) and convert to date range
        try:
            year, month_num = map(int, month.split("-"))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        
        start_date = f"{year}-{month_num:02d}-01"
        if month_num == 12:
            end_year = year + 1
            end_month = 1
        else:
            end_year = year
            end_month = month_num + 1
        end_date = f"{end_year}-{end_month:02d}-01"
        
        # Get roster data
        export_data = get_roster_data_for_export(start_date, end_date, current_user)
        
        if not export_data:
            raise HTTPException(status_code=404, detail="No roster data found for the specified month")
        
        # Create PDF in memory
        output = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        # Define styles
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Create story (content)
        story = []
        
        # Add title
        year_month = datetime.strptime(month + "-01", "%Y-%m-%d")
        month_name = year_month.strftime("%B %Y")
        title = Paragraph(f"Roster Export - {month_name}", title_style)
        story.append(title)
        
        # Prepare data for table
        if export_data:
            # Get headers
            headers = list(export_data[0].keys())
            
            # Create table data
            table_data = [headers]  # Header row
            
            for entry in export_data:
                row = [str(entry.get(key, '')) for key in headers]
                table_data.append(row)
            
            # Create table
            from reportlab.platypus import Table, TableStyle
            from reportlab.lib import colors
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        # Create filename
        filename = f"roster_export_{month.replace('-', '_')}.pdf"
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException as he:
        # Re-raise HTTP exceptions (like 404) without wrapping them
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# NEW DATE RANGE EXPORT ENDPOINTS
@app.get("/api/export/range/csv")
async def export_csv_range(start_date: str, end_date: str, current_user: dict = Depends(get_current_user)):
    """Export roster data as CSV file for custom date range"""
    
    try:
        # Validate date format
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
        # Get roster data
        export_data = get_roster_data_for_export(start_date, end_date, current_user)
        
        if not export_data:
            raise HTTPException(status_code=404, detail="No roster data found for the specified date range")
        
        # Create CSV in memory
        output = io.StringIO()
        if export_data:
            fieldnames = export_data[0].keys()
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(export_data)
        
        # Get CSV content
        csv_content = output.getvalue()
        output.close()
        
        # Create filename
        filename = f"roster_export_{start_date}_to_{end_date}.csv"
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating CSV: {str(e)}")

@app.get("/api/export/range/excel")
async def export_excel_range(start_date: str, end_date: str, current_user: dict = Depends(get_current_user)):
    """Export roster data as Excel file for custom date range"""
    
    try:
        # Validate date format
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
        # Get roster data
        export_data = get_roster_data_for_export(start_date, end_date, current_user)
        
        if not export_data:
            raise HTTPException(status_code=404, detail="No roster data found for the specified date range")
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        # Create workbook and worksheet
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Convert data to DataFrame
            df = pd.DataFrame(export_data)
            
            # Write to Excel
            df.to_excel(writer, sheet_name='Roster Export', index=False)
            
            # Get workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Roster Export']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Create filename
        filename = f"roster_export_{start_date}_to_{end_date}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")

@app.get("/api/export/range/pdf")
async def export_pdf_range(start_date: str, end_date: str, current_user: dict = Depends(get_current_user)):
    """Export roster data as PDF file for custom date range"""
    
    try:
        # Validate date format
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
        # Get roster data
        export_data = get_roster_data_for_export(start_date, end_date, current_user)
        
        if not export_data:
            raise HTTPException(status_code=404, detail="No roster data found for the specified date range")
        
        # Create PDF in memory
        output = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        # Define styles
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Create story (content)
        story = []
        
        # Add title
        title_text = f"Roster Export - {start_dt.strftime('%B %d, %Y')} to {end_dt.strftime('%B %d, %Y')}"
        title = Paragraph(title_text, title_style)
        story.append(title)
        
        # Prepare data for table
        if export_data:
            # Get headers
            headers = list(export_data[0].keys())
            
            # Create table data
            table_data = [headers]  # Header row
            
            for entry in export_data:
                row = [str(entry.get(key, '')) for key in headers]
                table_data.append(row)
            
            # Create table
            from reportlab.platypus import Table, TableStyle
            from reportlab.lib import colors
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        # Create filename
        filename = f"roster_export_{start_date}_to_{end_date}.pdf"
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# Initialize database with default admin user if not exists
def initialize_admin():
    """Initialize default admin user if not exists"""
    admin_user = db.users.find_one({"username": "Admin"})
    if not admin_user:
        print("Creating default admin user...")
        default_admin = User(
            id=str(uuid.uuid4()),
            username="Admin",
            pin_hash=hash_pin("0000"),  # Default PIN: 0000 (user can change this)
            role=UserRole.ADMIN,
            email="admin@company.com",
            first_name="System",
            last_name="Administrator",
            created_at=datetime.utcnow(),
            is_active=True,
            is_first_login=False  # Admin doesn't need to change PIN immediately
        )
        db.users.insert_one(default_admin.dict())
        print("✅ Default admin user created: Username=Admin, PIN=0000")
    else:
        print("✅ Admin user already exists - preserving existing PIN")

# Initialize admin on startup
initialize_admin()

# Initialize sample client data
initialize_sample_client()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)